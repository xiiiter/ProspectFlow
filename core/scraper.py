"""
Coleta leads no Google Maps e salva os resultados em planilhas.
"""

import re
import time
import json
import signal
import random
import sys
from pathlib import Path
from datetime import datetime

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Instale: pip install playwright && playwright install chromium")
    raise

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale: pip install openpyxl")
    raise

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

# Anti-detecção: User Agents rotativos
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
]


def delay_humano(minimo: float = 1.5, maximo: float = 4.0):
    """Delay humanizado entre ações."""
    time.sleep(random.uniform(minimo, maximo))


# Estado global para Ctrl+C / GUI
_GUI_MODE = False

_g = {
    "leads":        [],
    "arquivo":      "",
    "interrompido": False,
}


def _salvar_emergencia():
    leads   = _g["leads"]
    arquivo = _g["arquivo"]
    if not leads:
        return
    if arquivo:
        try:
            salvar_xlsx(leads, arquivo)
            print(f"{len(leads)} leads salvos em: {arquivo}")
            return
        except Exception as e:
            print(f"Falha xlsx ({e}), tentando JSON...")
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = arquivo.replace(".xlsx", f"_backup_{ts}.json") if arquivo else f"leads_backup_{ts}.json"
    try:
        with open(backup, "w", encoding="utf-8") as f:
            json.dump(leads, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        print(f"Falha ao salvar: {e}")


def _handler_ctrl_c(sig, frame):
    _g["interrompido"] = True
    _salvar_emergencia()
    sys.exit(0)


# CNPJ

CNPJ_VAZIO = {
    "cnpj": "", "razao_social": "", "situacao_cnpj": "",
    "atividade_principal": "", "data_abertura": "",
    "capital_social": "", "natureza_juridica": "",
    "telefone_cnpj": "", "nome_dono": "", "socios": "",
}


def _fmt_cnpj(n: str) -> str:
    n = re.sub(r"\D", "", n)
    return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:]}" if len(n) == 14 else n


def _brasilapi(cnpj_num: str) -> dict:
    try:
        r = requests.get(
            f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_num}",
            timeout=8, headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code == 200:
            d   = r.json()
            cap = d.get("capital_social", 0) or 0
            qsa = d.get("qsa", []) or []
            nomes = [s.get("nome_socio", "") or s.get("nome", "") for s in qsa]
            nomes = [n for n in nomes if n]
            return {
                "cnpj":                _fmt_cnpj(cnpj_num),
                "razao_social":        d.get("razao_social", ""),
                "situacao_cnpj":       d.get("descricao_situacao_cadastral", ""),
                "atividade_principal": d.get("cnae_fiscal_descricao", ""),
                "data_abertura":       d.get("data_inicio_atividade", ""),
                "capital_social":      f"R$ {cap:,.2f}".replace(",","X").replace(".",",").replace("X","."),
                "natureza_juridica":   d.get("descricao_natureza_juridica", ""),
                "telefone_cnpj":       "",
                "nome_dono":           nomes[0] if nomes else "",
                "socios":              " | ".join(nomes[:5]),
            }
    except Exception:
        pass
    return {}


def _cnpjws(cnpj_num: str) -> dict:
    try:
        r = requests.get(
            f"https://publica.cnpj.ws/cnpj/{cnpj_num}",
            timeout=10, headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code == 200:
            d = r.json()
            tel_raw = d.get("ddd_telefone_1", "") or d.get("ddd_telefone_2", "") or ""
            tel = re.sub(r"\D", "", tel_raw)
            qsa = d.get("qsa", []) or []
            nomes = [
                s.get("nome_socio", "") or s.get("nome_responsavel", "")
                for s in qsa
            ]
            nomes = [n for n in nomes if n]
            cap_raw = d.get("capital_social", "") or ""
            try:
                cap = float(str(cap_raw).replace(",", "."))
                cap_fmt = f"R$ {cap:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            except Exception:
                cap_fmt = str(cap_raw)
            return {
                "cnpj":                _fmt_cnpj(cnpj_num),
                "razao_social":        d.get("razao_social", ""),
                "situacao_cnpj":       (d.get("descricao_situacao_cadastral", "") or
                                        d.get("situacao_cadastral", "")),
                "atividade_principal": (d.get("cnae_fiscal_descricao", "") or
                                        d.get("descricao_cnae_principal", "")),
                "data_abertura":       d.get("data_inicio_atividade", ""),
                "capital_social":      cap_fmt,
                "natureza_juridica":   d.get("descricao_natureza_juridica", ""),
                "telefone_cnpj":       tel,
                "nome_dono":           nomes[0] if nomes else "",
                "socios":              " | ".join(nomes[:5]),
            }
    except Exception:
        pass
    return {}


def _casadosdados_search(nome_empresa: str, cidade: str = "") -> str:
    try:
        payload = {
            "query": {"razao_social": nome_empresa},
            "range_query": {},
            "extras": {"somente_mei": False, "excluir_sem_contato": False,
                       "com_email": False, "incluir_atividade_secundaria": False},
            "page": 1,
        }
        if cidade:
            mun = re.sub(r"\s+", " ", cidade.split("-")[0].split(",")[0]).strip().upper()
            payload["query"]["municipio"] = [mun]
        r = requests.post(
            "https://api.casadosdados.com.br/v2/public/cnpj/pesquisa",
            json=payload, timeout=10,
            headers={"User-Agent": "Mozilla/5.0", "Content-Type": "application/json",
                     "Accept": "application/json"},
        )
        if r.status_code == 200:
            items = r.json().get("data", {}).get("cnpj", [])
            if items:
                n = re.sub(r"\D", "", items[0].get("cnpj", "") or "")
                return n if len(n) == 14 else ""
    except Exception:
        pass
    return ""


def _receitaws(cnpj_num: str) -> dict:
    try:
        r = requests.get(
            f"https://www.receitaws.com.br/v1/cnpj/{cnpj_num}",
            timeout=12, headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code == 200:
            d = r.json()
            if d.get("status") != "ERROR":
                tel = re.sub(r"\D", "", d.get("telefone", "") or "")
                qsa = d.get("qsa", []) or []
                nomes = [s.get("nome", "") for s in qsa if s.get("nome")]
                cap_raw = d.get("capital_social", "") or ""
                try:
                    cap = float(re.sub(r"[^\d,]", "", cap_raw).replace(",", "."))
                    cap_fmt = f"R$ {cap:,.2f}".replace(",","X").replace(".",",").replace("X",".")
                except Exception:
                    cap_fmt = cap_raw
                return {
                    "cnpj":                _fmt_cnpj(cnpj_num),
                    "razao_social":        d.get("nome", ""),
                    "situacao_cnpj":       d.get("situacao", ""),
                    "atividade_principal": (d.get("atividades_principais") or [{}])[0].get("text", ""),
                    "data_abertura":       d.get("abertura", ""),
                    "capital_social":      cap_fmt,
                    "natureza_juridica":   d.get("natureza_juridica", ""),
                    "telefone_cnpj":       tel,
                    "nome_dono":           nomes[0] if nomes else "",
                    "socios":              " | ".join(nomes[:5]),
                }
    except Exception:
        pass
    return {}


def buscar_cnpj(nome_empresa: str, cidade: str = "") -> dict:
    """Busca CNPJ pelo nome da empresa com múltiplas estratégias."""
    if not REQUESTS_OK:
        return CNPJ_VAZIO.copy()

    hdrs = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    # Estratégia 1: Casa dos Dados
    cnpj_num = _casadosdados_search(nome_empresa, cidade)
    if cnpj_num:
        dados = _cnpjws(cnpj_num)
        if not dados:
            dados = _brasilapi(cnpj_num)
        if dados:
            return dados
        res = CNPJ_VAZIO.copy()
        res["cnpj"] = _fmt_cnpj(cnpj_num)
        return res

    # Estratégia 2: cnpj.biz scraping
    try:
        nome_clean = re.sub(r"[^\w\s]", " ", nome_empresa).strip()
        q2 = requests.utils.quote(nome_clean)
        r2 = requests.get(
            f"https://www.cnpj.biz/pesquisa/{q2}",
            timeout=10, headers={**hdrs, "Accept-Language": "pt-BR,pt;q=0.9"},
        )
        if r2.status_code == 200:
            cnpjs = re.findall(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", r2.text)
            if cnpjs:
                n = re.sub(r"\D", "", cnpjs[0])
                dados = _cnpjws(n) or _brasilapi(n)
                if dados:
                    return dados
                res = CNPJ_VAZIO.copy()
                res["cnpj"] = cnpjs[0]
                return res
    except Exception:
        pass

    # Estratégia 3: open.cnpja.com
    try:
        q3 = f"{nome_empresa} {cidade}".strip()
        r3 = requests.get(
            f"https://open.cnpja.com/office/search?q={requests.utils.quote(q3)}&limit=3",
            timeout=8, headers=hdrs,
        )
        if r3.status_code == 200:
            data    = r3.json()
            offices = data if isinstance(data, list) else data.get("offices", [])
            if not offices and isinstance(data, dict):
                offices = data.get("data", [])
            if offices:
                first     = offices[0]
                cnpj_num2 = re.sub(r"\D", "", first.get("taxId", "") or "")
                if len(cnpj_num2) == 14:
                    dados = _cnpjws(cnpj_num2) or _brasilapi(cnpj_num2)
                    if dados:
                        return dados
                    res = CNPJ_VAZIO.copy()
                    res["cnpj"] = _fmt_cnpj(cnpj_num2)
                    company = first.get("company", {}) or {}
                    res["razao_social"]  = company.get("name", "") or first.get("name", "")
                    status               = first.get("status", {}) or {}
                    res["situacao_cnpj"] = status.get("text", "")
                    res["data_abertura"] = first.get("founded", "")
                    return res
    except Exception:
        pass

    return CNPJ_VAZIO.copy()


# WhatsApp
def formatar_whatsapp(telefone: str) -> str:
    if not telefone:
        return ""
    nums = re.sub(r"\D", "", telefone)
    if nums.startswith("0"):
        nums = nums[1:]
    if len(nums) in (10, 11):
        nums = "55" + nums
    return f"https://wa.me/{nums}" if len(nums) >= 12 else ""


# Extratores de links (exclui feed lateral)
def _links_excluindo_feed(page, selector: str) -> list:
    try:
        todos   = page.locator(selector).all()
        no_feed = page.locator(f'div[role="feed"] {selector}').all()
        feed_hrefs = set()
        for el in no_feed:
            try:
                h = el.get_attribute("href", timeout=400)
                if h:
                    feed_hrefs.add(h)
            except Exception:
                pass
        result = []
        for el in todos:
            try:
                h = el.get_attribute("href", timeout=400)
                if h not in feed_hrefs:
                    result.append(el)
            except Exception:
                pass
        return result
    except Exception:
        return []


def extrair_instagram(page) -> str:
    for link in _links_excluindo_feed(page, 'a[href*="instagram.com"]')[:5]:
        try:
            href = link.get_attribute("href", timeout=800)
            if href and "instagram.com" in href:
                m = re.search(r"instagram\.com/([^/?#\s]+)", href)
                if m:
                    h = m.group(1)
                    if h.lower() not in ("p", "explore", "accounts", "tv", "reel", "stories"):
                        return f"@{h}"
        except Exception:
            pass
    return ""


def extrair_facebook(page) -> str:
    for link in _links_excluindo_feed(page, 'a[href*="facebook.com"]')[:5]:
        try:
            href = link.get_attribute("href", timeout=800)
            if href and "facebook.com" in href:
                m = re.search(r"facebook\.com/([^/?#\s]+)", href)
                if m and m.group(1).lower() not in (
                    "share", "sharer", "dialog", "pages", "login", "home.php"
                ):
                    return href
        except Exception:
            pass
    return ""


def extrair_whatsapp_pagina(page) -> str:
    for link in _links_excluindo_feed(page, 'a[href*="wa.me"], a[href*="whatsapp.com/send"]')[:5]:
        try:
            href = link.get_attribute("href", timeout=800)
            if href and "wa.me" in href:
                return href
        except Exception:
            pass
    return ""


def extrair_categoria(page) -> str:
    for sel in [
        'button[jsaction*="category"]',
        "[data-attrid='subtitle']",
        "span.DkEaL",
        "div.LBgpqf button",
        "button.DkEaL",
    ]:
        try:
            t = page.locator(sel).first.inner_text(timeout=1200).strip()
            if t and len(t) < 70:
                return t
        except Exception:
            pass
    return ""


def extrair_horarios(page) -> str:
    for sel in [
        'div[aria-label*="Horário"]',
        'div[aria-label*="Opening hours"]',
        'div[aria-label*="Hours"]',
        "table.WgFkxc",
        "div.t39EBf",
    ]:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1200):
                t      = el.inner_text(timeout=1200)
                linhas = [l.strip() for l in t.split("\n") if l.strip()]
                return " | ".join(linhas[:7])[:200]
        except Exception:
            pass
    return ""


def extrair_email_website(url: str) -> str:
    if not url or not REQUESTS_OK:
        return ""
    skip = ("facebook.com","instagram.com","twitter.com","tiktok.com",
            "youtube.com","linkedin.com","maps.google.com","wa.me")
    if any(d in url for d in skip):
        return ""
    try:
        resp   = requests.get(url, timeout=7, headers={"User-Agent":"Mozilla/5.0"}, allow_redirects=True)
        emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", resp.text)
        noise  = {"sentry","example","test","schema","jquery","noreply","no-reply","@2x"}
        validos = [e for e in emails
                   if not any(n in e.lower() for n in noise) and 6 < len(e) < 80]
        return validos[0] if validos else ""
    except Exception:
        return ""


# Coleta URLs do feed de resultados
def _coletar_urls_feed(page, max_needed: int) -> list:
    urls   = []
    vistos = set()

    for sel in [
        'div[role="feed"] a[href*="/maps/place/"]',
        'a[href*="/maps/place/"]',
    ]:
        try:
            elementos = page.locator(sel).all()
            for el in elementos:
                try:
                    href = el.get_attribute("href", timeout=500)
                    if href and href.startswith("http") and "/maps/place/" in href and href not in vistos:
                        vistos.add(href)
                        urls.append(href)
                        if len(urls) >= max_needed * 2:
                            break
                except Exception:
                    pass
            if urls:
                break
        except Exception:
            pass

    return urls


# Extração de dados de um painel de negócio
def _extrair_negocio(page, nicho: str, cidade: str, buscar_cnpj_flag: bool, buscar_dono: bool = False) -> dict:
    try:
        time.sleep(0.3)

        nome = ""
        for sel_h1 in ["h1", "h1.DUwDvf", "h1.fontHeadlineLarge"]:
            try:
                nome = page.locator(sel_h1).first.inner_text(timeout=3000).strip()
                if nome:
                    break
            except Exception:
                pass
        if not nome:
            return None

        categoria = extrair_categoria(page)

        # Rating / reviews
        rating, num_reviews = "", ""
        try:
            for sel_rat in [
                'div[jsaction*="pane.rating"]',
                'span[aria-label*="estrelas"]',
                'span[aria-label*="stars"]',
            ]:
                try:
                    rt = page.locator(sel_rat).first.inner_text(timeout=2000)
                    m  = re.search(r"([\d,\.]+)\s*\(([\d,\.]+)", rt)
                    if m:
                        rating      = m.group(1).replace(",", ".")
                        num_reviews = re.sub(r"\D", "", m.group(2))
                        break
                except Exception:
                    pass
            if not rating:
                try:
                    lbl = page.locator('span[aria-label*="estrelas"], span[aria-label*="stars"]').first.get_attribute("aria-label", timeout=2000) or ""
                    m2  = re.search(r"([\d,\.]+)", lbl)
                    if m2:
                        rating = m2.group(1).replace(",", ".")
                except Exception:
                    pass
        except Exception:
            pass

        # Endereço
        endereco = ""
        for sel_end in [
            'button[data-item-id="address"]',
            '[data-item-id="address"] span.Io6YTe',
            'div[data-item-id="address"]',
        ]:
            try:
                endereco = page.locator(sel_end).first.inner_text(timeout=2000).strip()
                if endereco:
                    break
            except Exception:
                pass

        # Telefone
        telefone = ""
        for sel_tel in [
            'button[data-item-id*="phone"]',
            '[data-item-id*="phone"] span.Io6YTe',
        ]:
            try:
                telefone = page.locator(sel_tel).first.inner_text(timeout=2000).strip()
                if telefone:
                    break
            except Exception:
                pass

        # WhatsApp
        whatsapp = extrair_whatsapp_pagina(page) or formatar_whatsapp(telefone)

        # Website
        website, site_tipo = "", "sem_site"
        for sel_site in [
            'a[data-item-id="authority"]',
            'a[href^="http"][data-item-id]',
        ]:
            try:
                href = page.locator(sel_site).first.get_attribute("href", timeout=2000)
                if href and "google.com" not in href:
                    website   = href
                    site_tipo = ("apenas_social"
                                 if ("facebook.com" in href or "instagram.com" in href)
                                 else "tem_site")
                    break
            except Exception:
                pass

        instagram = extrair_instagram(page)
        facebook  = extrair_facebook(page)
        horario   = extrair_horarios(page)
        maps_url  = page.url

        plus_code = ""
        try:
            plus_code = page.locator('[data-item-id="oloc"] span.Io6YTe').first.inner_text(timeout=1000).strip()
        except Exception:
            pass

        email = ""
        if website and site_tipo == "tem_site" and REQUESTS_OK:
            email = extrair_email_website(website)

        dados_cnpj = CNPJ_VAZIO.copy()
        if buscar_cnpj_flag and REQUESTS_OK:
            dados_cnpj = buscar_cnpj(nome, cidade)
            if buscar_dono and dados_cnpj.get("cnpj") and not dados_cnpj.get("telefone_cnpj"):
                cnpj_num = re.sub(r"\D", "", dados_cnpj["cnpj"])
                extra = _cnpjws(cnpj_num)
                if not extra:
                    extra = _receitaws(cnpj_num)
                for k in ("telefone_cnpj", "nome_dono", "socios"):
                    if extra.get(k) and not dados_cnpj.get(k):
                        dados_cnpj[k] = extra[k]

        telefone_dono = dados_cnpj.get("telefone_cnpj", "")
        if telefone_dono and len(telefone_dono) >= 10:
            d = telefone_dono
            if len(d) == 10:
                telefone_dono = f"({d[:2]}) {d[2:6]}-{d[6:]}"
            elif len(d) == 11:
                telefone_dono = f"({d[:2]}) {d[2:7]}-{d[7:]}"

        tem_site = site_tipo != "sem_site"
        return {
            "nome":         nome,
            "nicho":        nicho,
            "cidade":       cidade,
            "categoria":    categoria,
            "rating":       rating,
            "num_reviews":  num_reviews,
            "prioridade":   _calcular_prioridade(tem_site, site_tipo, rating),
            "site_tipo":    site_tipo,
            "website":      website,
            "email":        email,
            "telefone":     telefone,
            "whatsapp":     whatsapp,
            "instagram":    instagram,
            "facebook":     facebook,
            "endereco":     endereco,
            "plus_code":    plus_code,
            "horario":      horario,
            "maps_url":     maps_url,
            "cnpj":                dados_cnpj.get("cnpj", ""),
            "razao_social":        dados_cnpj.get("razao_social", ""),
            "situacao_cnpj":       dados_cnpj.get("situacao_cnpj", ""),
            "atividade_principal": dados_cnpj.get("atividade_principal", ""),
            "data_abertura":       dados_cnpj.get("data_abertura", ""),
            "capital_social":      dados_cnpj.get("capital_social", ""),
            "natureza_juridica":   dados_cnpj.get("natureza_juridica", ""),
            "telefone_dono":       telefone_dono,
            "nome_dono":           dados_cnpj.get("nome_dono", ""),
            "socios":              dados_cnpj.get("socios", ""),
            "status":          "novo",
            "data_encontrado": datetime.now().strftime("%Y-%m-%d"),
            "notas":           "",
            "email_subject":   "",
            "email_body":      "",
            "preco":           "",
            "email_gerado":    "",
        }
    except Exception:
        return None


def _calcular_prioridade(tem_site, site_tipo, rating) -> str:
    if not tem_site or site_tipo == "sem_site":
        return "Alta"
    if site_tipo == "apenas_social":
        return "Media"
    return "Baixa"


def _classificar(lead) -> str:
    return {
        "sem_site":     "Sem site",
        "apenas_social":"So social",
        "tem_site":     "Tem site",
    }.get(lead.get("site_tipo", ""), "?")


def _icones(lead) -> str:
    return (
        ("WA" if lead.get("whatsapp")  else "·") + " " +
        ("CNPJ" if lead.get("cnpj")    else "·") + " " +
        ("Email" if lead.get("email")  else "·") + " " +
        ("Insta" if lead.get("instagram") else "·")
    )


# Scraper principal
def scrape_google_maps(
    nicho: str,
    cidade: str,
    max_leads: int = 30,
    buscar_cnpj_flag: bool = True,
    buscar_dono: bool = False,
    salvar_a_cada: int = 5,
    on_lead=None,
    on_captcha=None,
    captcha_evento=None,
) -> list:
    """
    Scrapa Google Maps extraindo o máximo de dados por negócio.
    Usa navegação por URL direta (mais confiável que clique em card).

    on_captcha: callback chamado quando CAPTCHA é detectado
    captcha_evento: threading.Event para aguardar resolução
    """
    tem_acento = any(c in (cidade + nicho) for c in "áéíóúãõàèìòùâêîôûçÁÉÍÓÚÃÕÀÈÌÒÙÂÊÎÔÛÇ")
    sep        = "em" if tem_acento else "in"
    query      = f"{nicho} {sep} {cidade}"
    leads      = []
    vistos     = set()

    # Seleciona user agent aleatório
    user_agent = random.choice(USER_AGENTS)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=user_agent,
            locale="pt-BR" if tem_acento else "en-US",
        )
        page = context.new_page()
        page.set_default_timeout(60000)

        # Abre página de busca
        url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
        except Exception:
            page.goto(url, timeout=60000)

        # Detecta CAPTCHA antes de prosseguir
        _checar_captcha(page, on_captcha, captcha_evento)

        try:
            page.wait_for_selector('div[role="feed"]', timeout=20000)
        except Exception:
            pass

        time.sleep(2)

        # Aceita cookies/LGPD
        for sel in ['button[aria-label*="Accept"]', 'button[aria-label*="Aceitar"]',
                    'form[action*="consent"] button']:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=2000):
                    btn.click()
                    time.sleep(1)
                    break
            except Exception:
                pass

        # Scroll para carregar mais resultados
        n_scrolls   = max(10, max_leads // 2)
        stall_count = 0
        prev_count  = 0
        try:
            feed = page.locator('div[role="feed"]')
            for s in range(n_scrolls):
                if _g["interrompido"]:
                    break
                feed.evaluate("el => el.scrollBy(0, 1800)")
                time.sleep(random.uniform(1.2, 1.8))
                if s > 5:
                    hrefs_tmp     = _coletar_urls_feed(page, max_leads)
                    current_count = len(hrefs_tmp)
                    if current_count >= max_leads:
                        break
                    if current_count == prev_count:
                        stall_count += 1
                        if stall_count >= 3:
                            break
                    else:
                        stall_count = 0
                    prev_count = current_count
        except Exception:
            for _ in range(n_scrolls):
                page.keyboard.press("End")
                time.sleep(1.8)

        all_hrefs = _coletar_urls_feed(page, max_leads)
        total     = len(all_hrefs)

        if not all_hrefs:
            browser.close()
            return leads

        # Processa cada URL individualmente
        for i, href in enumerate(all_hrefs):
            if _g["interrompido"]:
                break

            # Verifica CAPTCHA antes de cada lead
            _checar_captcha(page, on_captcha, captcha_evento)

            try:
                try:
                    page.goto(href, wait_until="domcontentloaded", timeout=30000)
                except Exception:
                    try:
                        page.goto(href, timeout=30000)
                    except Exception:
                        continue

                try:
                    page.wait_for_selector("h1", timeout=12000)
                except Exception:
                    continue

                time.sleep(random.uniform(0.8, 1.5))

                lead = _extrair_negocio(page, nicho, cidade, buscar_cnpj_flag, buscar_dono)
                if not lead:
                    continue

                # Deduplicação
                chave = (lead["nome"].lower().strip(), lead["endereco"][:25].lower())
                if chave in vistos:
                    continue
                vistos.add(chave)

                leads.append(lead)
                _g["leads"] = leads

                if on_lead:
                    try:
                        on_lead(lead)
                    except Exception:
                        pass

                # Auto-save
                if _g["arquivo"] and len(leads) % salvar_a_cada == 0:
                    try:
                        salvar_xlsx(leads, _g["arquivo"])
                    except Exception:
                        pass

            except Exception:
                continue

        browser.close()

    return leads


def _checar_captcha(page, on_captcha=None, captcha_evento=None):
    """Detecta CAPTCHA e aguarda resolução se on_captcha/captcha_evento fornecidos."""
    try:
        content = page.content().lower()
        if any(x in content for x in ['unusual traffic', 'recaptcha', 'captcha', 'unusual traffic from']):
            if on_captcha:
                on_captcha()
                if captcha_evento:
                    captcha_evento.wait(timeout=300)
                    captcha_evento.clear()
    except Exception:
        pass


# Retype Turbo
_SINONIMOS: dict = {
    "dentista":          ["dentista", "clínica odontológica", "consultório dentário", "odontologista"],
    "clínica odontológica": ["clínica odontológica", "dentista", "odontologia"],
    "ortodontista":      ["ortodontista", "aparelho dentário", "ortodontia"],
    "fisioterapeuta":    ["fisioterapeuta", "clínica de fisioterapia", "fisioterapia"],
    "psicólogo":         ["psicólogo", "psicóloga", "clínica de psicologia"],
    "nutricionista":     ["nutricionista", "consultório nutrição", "nutrição"],
    "médico":            ["médico", "clínica médica", "consultório médico"],
    "dermatologista":    ["dermatologista", "clínica de dermatologia", "dermatologia"],
    "veterinário":       ["veterinário", "clínica veterinária", "pet care"],
    "pet shop":          ["pet shop", "petshop", "loja de animais"],
    "barbearia":         ["barbearia", "barbershop", "barber", "barbeiro"],
    "salão de beleza":   ["salão de beleza", "cabeleireiro", "salão feminino"],
    "estética":          ["estética", "clínica estética", "estéticista"],
    "academia":          ["academia", "gym", "musculação", "crossfit", "fitness"],
    "restaurante":       ["restaurante", "lanchonete", "comida caseira", "gastronomia"],
    "pizzaria":          ["pizzaria", "pizza delivery", "pizza artesanal"],
    "hamburgueria":      ["hamburgueria", "hamburguer artesanal", "burger"],
    "padaria":           ["padaria", "confeitaria", "panificadora"],
    "cafeteria":         ["cafeteria", "café", "coffee shop"],
    "advogado":          ["advogado", "advocacia", "escritório de advocacia", "advogada"],
    "contador":          ["contador", "contabilidade", "escritório contábil"],
    "imobiliária":       ["imobiliária", "corretor de imóveis", "agência imobiliária"],
    "eletricista":       ["eletricista", "instalações elétricas", "elétrica"],
    "mecânica":          ["mecânica", "oficina mecânica", "mecânico"],
    "dentist":           ["dentist", "dental clinic", "dental office"],
    "barbershop":        ["barbershop", "barber", "hair salon"],
    "gym":               ["gym", "fitness center", "crossfit"],
    "restaurant":        ["restaurant", "diner", "food"],
    "lawyer":            ["lawyer", "law office", "attorney"],
    "accountant":        ["accountant", "accounting firm", "CPA"],
}


def _get_sinonimos(nicho: str) -> list:
    nl = nicho.lower().strip()
    for k, v in _SINONIMOS.items():
        if k == nl or nl in k or k in nl:
            seen = set()
            result = []
            for s in v:
                if s not in seen:
                    seen.add(s)
                    result.append(s)
            return result
    return [nicho]


def scrape_turbo_retype(
    nicho: str,
    cidade: str,
    meta_alta: int = 50,
    max_por_query: int = 40,
    buscar_cnpj_flag: bool = True,
    buscar_dono: bool = False,
    salvar_a_cada: int = 999,
    on_lead=None,
    on_captcha=None,
    captcha_evento=None,
) -> list:
    """
    Retype Turbo: coleta leads via múltiplas queries (sinônimos) até atingir
    `meta_alta` leads de Alta prioridade.
    """
    sinonimos = _get_sinonimos(nicho)
    todos_leads: list = []
    vistos_global: set = set()
    alta_count = 0

    for variante in sinonimos:
        if alta_count >= meta_alta:
            break

        restante = meta_alta - alta_count
        buscar = max(max_por_query, restante * 3)

        def _on_lead_wrapper(lead: dict):
            nonlocal alta_count
            if "Alta" in lead.get("prioridade", ""):
                alta_count += 1
            if on_lead:
                try:
                    on_lead(lead)
                except Exception:
                    pass

        novos = scrape_google_maps(
            nicho=variante,
            cidade=cidade,
            max_leads=buscar,
            buscar_cnpj_flag=buscar_cnpj_flag,
            buscar_dono=buscar_dono,
            salvar_a_cada=salvar_a_cada,
            on_lead=_on_lead_wrapper,
            on_captcha=on_captcha,
            captcha_evento=captcha_evento,
        )

        for lead in novos:
            chave = (lead.get("nome", "").lower().strip(), (lead.get("endereco") or "")[:25].lower())
            if chave not in vistos_global:
                vistos_global.add(chave)
                todos_leads.append(lead)

    return todos_leads


# Colunas da planilha
HEADERS = [
    ("Nome","nome"),("Nicho","nicho"),("Cidade","cidade"),("Categoria","categoria"),
    ("Rating","rating"),("Reviews","num_reviews"),("Prioridade","prioridade"),
    ("Site Tipo","site_tipo"),("Website","website"),("Email","email"),
    ("Telefone","telefone"),("WhatsApp","whatsapp"),("Instagram","instagram"),
    ("Facebook","facebook"),("Endereço","endereco"),("Plus Code","plus_code"),
    ("Horário","horario"),("Maps URL","maps_url"),
    ("CNPJ","cnpj"),("Razão Social","razao_social"),
    ("Situação CNPJ","situacao_cnpj"),("Atividade Principal","atividade_principal"),
    ("Data Abertura","data_abertura"),("Capital Social","capital_social"),
    ("Tel. Dono (RF)","telefone_dono"),("Nome Dono","nome_dono"),("Sócios","socios"),
    ("Natureza Jurídica","natureza_juridica"),
    ("Status","status"),("Data","data_encontrado"),("Notas","notas"),
    ("Email Subject","email_subject"),("Email Body","email_body"),
    ("Preço","preco"),("Email Gerado","email_gerado"),
]

COL_WIDTHS = {
    "nome":34,"nicho":16,"cidade":18,"categoria":22,"rating":8,"num_reviews":9,
    "prioridade":13,"site_tipo":14,"website":38,"email":32,"telefone":18,
    "whatsapp":34,"instagram":22,"facebook":36,"endereco":44,"plus_code":18,
    "horario":44,"maps_url":36,"cnpj":22,"razao_social":36,"situacao_cnpj":16,
    "atividade_principal":40,"data_abertura":14,"capital_social":18,
    "natureza_juridica":32,
    "telefone_dono":18,"nome_dono":34,"socios":50,
    "status":16,"data_encontrado":14,"notas":32,
    "email_subject":40,"email_body":60,"preco":12,"email_gerado":14,
}

URL_KEYS = {"website","whatsapp","facebook","maps_url"}


# Salvar Excel
def salvar_xlsx(leads: list, caminho: str = "data/leads/leads.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    hf = PatternFill("solid", start_color="1F4E79")
    ff = Font(bold=True, color="FFFFFF", size=11)
    for col, (h, _) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = ff
        c.alignment = Alignment(horizontal="center", vertical="center")

    fills = {
        "alta":  PatternFill("solid", start_color="FFE0E0"),
        "media": PatternFill("solid", start_color="FFF3CD"),
        "baixa": PatternFill("solid", start_color="E8F5E9"),
        "cinza": PatternFill("solid", start_color="F8F8F8"),
    }

    for ri, lead in enumerate(leads, 2):
        prio = lead.get("prioridade", "")
        fill = (fills["alta"]  if "Alta"  in prio else
                fills["media"] if "Media" in prio else
                fills["baixa"] if "Baixa" in prio else fills["cinza"])
        for ci, (_, key) in enumerate(HEADERS, 1):
            v = lead.get(key, "") or ""
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=(key == "email_body"))
            if key in URL_KEYS and v and str(v).startswith("http"):
                c.hyperlink = v
                c.font = Font(color="0563C1", underline="single")

    for ci, (_, k) in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(k, 20)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Aba Resumo
    ws2 = wb.create_sheet("Resumo")
    stats = [
        ("Total de leads",       "=COUNTA(Leads!A2:A99999)"),
        ("Alta prioridade",      '=COUNTIF(Leads!G2:G99999,"*Alta*")'),
        ("Media prioridade",     '=COUNTIF(Leads!G2:G99999,"*Media*")'),
        ("Baixa prioridade",     '=COUNTIF(Leads!G2:G99999,"*Baixa*")'),
        ("Sem site",             '=COUNTIF(Leads!H2:H99999,"sem_site")'),
        ("So social",            '=COUNTIF(Leads!H2:H99999,"apenas_social")'),
        ("Tem site",             '=COUNTIF(Leads!H2:H99999,"tem_site")'),
        ("Com WhatsApp",         '=COUNTIFS(Leads!L2:L99999,"http*")'),
        ("Com Email",            '=COUNTA(Leads!J2:J99999)-COUNTBLANK(Leads!J2:J99999)'),
        ("Com CNPJ",             '=COUNTA(Leads!S2:S99999)-COUNTBLANK(Leads!S2:S99999)'),
        ("Com Instagram",        '=COUNTA(Leads!M2:M99999)-COUNTBLANK(Leads!M2:M99999)'),
        ("Emails gerados",       '=COUNTIF(Leads!AF2:AF99999,"sim")'),
        ("Emails enviados",      '=COUNTIF(Leads!Z2:Z99999,"email_enviado")'),
        ("Responderam",          '=COUNTIF(Leads!Z2:Z99999,"respondeu")'),
        ("Fechados",             '=COUNTIF(Leads!Z2:Z99999,"fechado")'),
    ]
    hf2 = PatternFill("solid", start_color="1F4E79")
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hf2
    ws2["A1"] = "Metrica"
    ws2["B1"] = "Valor"
    for i, (lbl, fml) in enumerate(stats, 2):
        ws2.cell(i, 1, lbl).font = Font(bold=True)
        ws2.cell(i, 2, fml)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14

    Path(caminho).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
