#!/usr/bin/env python3
"""
core/whatsapp.py — Verificacao de WhatsApp + Auto Conversa

Combina:
  - Classificacao de numeros (fixo/celular)
  - Verificacao via WhatsApp Web (Playwright)
  - Deteccao de bot/mensagem automatica
  - Auto Conversa com Ollama (IA local)

Caminhos de dados:
  - enviados_log.json  -> data/logs/enviados_log.json
  - conversa_historico.json -> data/logs/conversa_historico.json
  - wa_profile/ -> wa_profile/ (raiz do projeto)

Instalacao:
    pip install playwright requests
    playwright install chromium

    Ollama rodando localmente:
    - Baixe em: https://ollama.ai/
    - Execute: ollama pull llama3
"""

import re
import sys
import time
import json
import signal
import hashlib
import argparse
import requests
from pathlib import Path
from datetime import datetime

# Caminhos de dados
_BASE_DIR = Path(__file__).parent.parent
LOG_ENVIADOS    = str(_BASE_DIR / "data" / "logs" / "enviados_log.json")
HISTORICO_FILE  = str(_BASE_DIR / "data" / "logs" / "conversa_historico.json")
AGENTE_CFG_FILE = str(_BASE_DIR / "data" / "logs" / "agente_config.json")
WA_PROFILE_DIR  = str(_BASE_DIR / "wa_profile")

# Configuracao Ollama
OLLAMA_URL    = "http://localhost:11434"
MODELO_PADRAO = "llama3"

# Padroes de mensagem automatica (bot)
_BOT_PATTERNS = [
    r"bem[- ]?vindo",
    r"boas[- ]?vindas",
    r"atendimento\s+(automatico|virtual|digital|eletronico)",
    r"assistente\s+virtual",
    r"chatbot",
    r"men[uo]\s*(de\s*)?(opcoes|atendimento|principal|servicos)",
    r"(pressione|digite|envie|escolha|selecione)\s+\d",
    r"\b[1-9]\s*[-\u2013.]\s*\w{3,}",
    r"\b(opcao|option)\s+\d",
    r"mensagem\s+automatica",
    r"resposta\s+automatica",
    r"este\s+e\s+um\s+(bot|robo|atendente\s+virtual)",
    r"\brobo\b",
    r"(fora|horario)\s+d[eo]\s+atendimento",
    r"aguarde\s+(um\s+momento|nosso\s+atendente|atendimento)",
    r"em\s+breve\s+(um\s+atendente|retornaremos)",
    r"para\s+(falar|continuar|agendar|saber|ver)\s+",
    r"(seu\s+)?(protocolo|ticket|chamado)\s*(numero|#|n[ou]\.?)?\s*\d+",
    r"numero\s+de\s+atendimento",
    r"obrigad[oa]\s+por\s+(entrar\s+em\s+contato|nos\s+contatar)",
    r"automated\s+(response|reply|message)",
    r"(press|type|send)\s+\d",
    r"out\s+of\s+(office|hours)",
    r"business\s+hours",
    r"virtual\s+assistant",
    r"thank\s+you\s+for\s+contacting",
    r"atendimento\s+via\s+whatsapp",
    r"horario\s+de\s+atendimento",
    r"nossa\s+equipe\s+(ira|vai|esta)",
    r"retornaremos\s+em\s+breve",
    r"seg(unda)?(\s*a\s*|-|\/)?\s*sex(ta)?",
]

_BOT_RE = [re.compile(p, re.IGNORECASE) for p in _BOT_PATTERNS]

_SEM_INTERESSE = [
    "nao tenho interesse", "nao tenho interesse",
    "nao preciso", "nao preciso",
    "nao quero", "nao quero",
    "nao me interessa", "nao me interessa",
    "remova", "remove meu", "me tire",
    "descadastre", "para de me mandar",
    "nao me contate", "nao me contate",
    "obrigado mas nao", "obrigado mas nao",
    "no thanks", "not interested",
    "ja tenho", "ja tenho",
    "nao preciso de site", "nao preciso de site",
]


def _eh_bot(texto: str) -> bool:
    """Retorna True se o texto parece mensagem automatica/bot."""
    if not texto or len(texto) < 8:
        return False
    return any(rx.search(texto) for rx in _BOT_RE)


def _eh_sem_interesse(texto: str) -> bool:
    t = texto.lower()
    return any(p in t for p in _SEM_INTERESSE)


def _hash_msg(texto: str) -> str:
    return hashlib.md5(texto.strip().encode()).hexdigest()[:12]


# Classificacao de numero brasileiro
def classificar_numero(telefone: str) -> str:
    """
    Classifica um numero de telefone brasileiro.
    Returns: "fixo" | "celular" | "desconhecido"
    """
    if not telefone:
        return "desconhecido"

    nums = re.sub(r"\D", "", str(telefone))

    if nums.startswith("55") and len(nums) > 12:
        nums = nums[2:]

    if nums.startswith("0") and len(nums) in (12, 13):
        nums = nums[1:]

    if len(nums) < 10 or len(nums) > 11:
        return "desconhecido"

    numero_local = nums[2:]

    if len(numero_local) == 9 and numero_local[0] == "9":
        return "celular"

    if len(numero_local) == 8 and numero_local[0] in "2345":
        return "fixo"

    if len(numero_local) == 8 and numero_local[0] in "678":
        return "fixo"

    return "desconhecido"


def formatar_numero_wa(telefone: str) -> str:
    """Formata o numero para uso no WhatsApp Web (+55XXXXXXXXXXX)."""
    nums = re.sub(r"\D", "", str(telefone))
    if nums.startswith("55") and len(nums) > 12:
        return nums
    if len(nums) <= 11:
        return "55" + nums
    return nums


# Persistencia
def carregar_historico() -> dict:
    p = Path(HISTORICO_FILE)
    if p.exists():
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def salvar_historico(historico: dict):
    Path(HISTORICO_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
        json.dump(historico, f, ensure_ascii=False, indent=2)


def carregar_autorizados() -> dict:
    """
    Carrega e retorna numeros autorizados do enviados_log.json.
    Returns: {numero_formatado: {nome, nicho, cidade, ts}}
    """
    p = Path(LOG_ENVIADOS)
    if not p.exists():
        return {}

    with open(p, encoding="utf-8") as f:
        data = json.load(f)

    wa_data = data.get("whatsapp", {})
    autorizados = {}

    for numero, info in wa_data.items():
        nums = re.sub(r"\D", "", numero)
        if not nums:
            continue
        if not nums.startswith("55"):
            nums = "55" + nums

        if isinstance(info, str):
            autorizados[nums] = {"nome": "", "nicho": "", "cidade": "", "ts": info}
        elif isinstance(info, dict):
            autorizados[nums] = {
                "nome":   info.get("nome", ""),
                "nicho":  info.get("nicho", ""),
                "cidade": info.get("cidade", ""),
                "ts":     info.get("ts", ""),
            }

    return autorizados


# Config do agente
_CFG_DEFAULTS = {
    "nome_agente":      "Consultor Digital",
    "empresa":          "Agencia de Sites",
    "objetivo":         "Agendar uma conversa rapida de 15-20 minutos",
    "servicos":         (
        "Sites profissionais personalizados e rapidos\n"
        "Aparecer no Google (SEO local)\n"
        "Funil de captacao de clientes online\n"
        "Landing pages para campanhas\n"
        "Gestao de presenca digital"
    ),
    "regras_extras":    "",
    "system_prompt_custom": "",
}


def carregar_agente_cfg() -> dict:
    p = Path(AGENTE_CFG_FILE)
    cfg = dict(_CFG_DEFAULTS)
    if p.exists():
        try:
            with open(p, encoding="utf-8") as f:
                cfg.update(json.load(f))
        except Exception:
            pass
    return cfg


def _montar_system_prompt(cfg: dict, nome: str, nicho: str, cidade: str) -> str:
    """Gera o system prompt a partir da config do agente."""
    if cfg.get("system_prompt_custom", "").strip():
        return cfg["system_prompt_custom"].format(
            nome=nome, nicho=nicho, cidade=cidade,
            nome_agente=cfg.get("nome_agente", ""),
            empresa=cfg.get("empresa", ""),
        )

    servicos_fmt = "\n".join(
        f"- {s.strip()}"
        for s in cfg.get("servicos", "").splitlines()
        if s.strip()
    )
    regras_extras = cfg.get("regras_extras", "").strip()
    regras_bloco  = f"\nREGRAS ADICIONAIS:\n{regras_extras}" if regras_extras else ""

    return (
        f"Voce e {cfg.get('nome_agente','Consultor Digital')}, representando "
        f"{cfg.get('empresa','a empresa')} — especializada em criacao de sites e "
        f"presenca digital para pequenas e medias empresas brasileiras.\n\n"
        f"CONTEXTO DO LEAD:\n"
        f"- Nome/Empresa: {nome}\n"
        f"- Segmento: {nicho}\n"
        f"- Cidade: {cidade}\n\n"
        f"OBJETIVO: {cfg.get('objetivo','Agendar uma conversa rapida')}\n\n"
        f"REGRAS ABSOLUTAS:\n"
        f"- Mensagens curtas: maximo 2-3 frases — e WhatsApp, nao email\n"
        f"- Nunca minta, exagere ou faca promessas que nao pode cumprir\n"
        f"- Jamais seja insistente — se o cliente disse nao, aceite graciosamente\n"
        f"- Use linguagem natural e calorosa, como uma pessoa real\n"
        f"- Adapte a abordagem ao segmento do cliente\n"
        f"- Se demonstrar interesse, proponha um horario para conversa breve\n"
        f"- Nunca mencione que voce e uma IA ou robo"
        f"{regras_bloco}\n\n"
        f"SERVICOS (mencione apenas quando relevante):\n{servicos_fmt}\n\n"
        f"Responda APENAS com o texto da mensagem, sem introducoes, aspas ou formatacao extra."
    )


# Ollama
def verificar_ollama(ollama_url: str, modelo: str) -> bool:
    """Verifica se Ollama esta rodando e o modelo esta disponivel."""
    try:
        resp = requests.get(f"{ollama_url}/api/tags", timeout=5)
        if resp.status_code != 200:
            return False

        modelos_disponiveis = [
            m["name"].split(":")[0]
            for m in resp.json().get("models", [])
        ]

        modelo_base = modelo.split(":")[0]
        if modelo_base not in modelos_disponiveis:
            if modelos_disponiveis:
                print(f"  Aviso: Modelo '{modelo}' nao encontrado.")
                print(f"     Disponiveis: {', '.join(modelos_disponiveis)}")
                print(f"     Execute: ollama pull {modelo}")
            else:
                print(f"  Aviso: Nenhum modelo instalado no Ollama.")
                print(f"     Execute: ollama pull {modelo}")
            return False

        return True

    except requests.exceptions.ConnectionError:
        return False
    except Exception as e:
        print(f"  Aviso: Erro verificando Ollama: {e}")
        return False


def gerar_resposta_ollama(
    ollama_url: str,
    modelo: str,
    numero: str,
    info: dict,
    historico: dict,
    nova_mensagem: str,
    agente_cfg: dict | None = None,
) -> str:
    """Gera resposta via Ollama com contexto do lead e historico da conversa."""
    cfg = agente_cfg or carregar_agente_cfg()
    sistema = _montar_system_prompt(
        cfg,
        nome=info.get("nome") or "cliente",
        nicho=info.get("nicho") or "negocio",
        cidade=info.get("cidade") or "",
    )

    msgs_api = [{"role": "system", "content": sistema}]

    hist_msgs = historico.get(numero, {}).get("mensagens", [])
    for msg in hist_msgs[-10:]:
        if msg.get("tipo") == "bot":
            continue
        role = "assistant" if msg["de"] == "nos" else "user"
        msgs_api.append({"role": role, "content": msg["texto"]})

    msgs_api.append({"role": "user", "content": nova_mensagem})

    try:
        resp = requests.post(
            f"{ollama_url}/api/chat",
            json={
                "model": modelo,
                "messages": msgs_api,
                "stream": False,
                "options": {
                    "temperature": 0.75,
                    "num_ctx": 4096,
                    "num_predict": 300,
                },
            },
            timeout=120,
        )

        if resp.status_code == 200:
            content = resp.json().get("message", {}).get("content", "").strip()
            content = content.strip('"\'').strip()
            content = re.sub(r"^(resposta|mensagem|reply)\s*:\s*", "", content, flags=re.IGNORECASE)
            return content
        else:
            print(f"  Ollama HTTP {resp.status_code}: {resp.text[:120]}")

    except requests.exceptions.Timeout:
        print("  Ollama timeout — modelo demorou demais, tente um modelo menor")
    except Exception as e:
        print(f"  Erro Ollama: {e}")

    return ""


# WhatsApp Web — Verificador
class WhatsAppVerificador:
    """
    Verifica numeros via WhatsApp Web usando Playwright.
    Usado para classificar se um numero tem WA ativo e detectar bots.
    """

    def __init__(self, profile_dir: str = WA_PROFILE_DIR, esperar_bot: int = 10):
        self.profile_dir = Path(profile_dir)
        self.esperar_bot = esperar_bot
        self._playwright = None
        self._context = None
        self._page = None

    def iniciar(self):
        """Abre WhatsApp Web com perfil persistente. QR so na 1a vez."""
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            print("Playwright nao instalado:")
            print("   pip install playwright && playwright install chromium")
            raise

        self._playwright = sync_playwright().__enter__()
        self.profile_dir.mkdir(parents=True, exist_ok=True)

        primeiro_uso = not (self.profile_dir / "Default" / "Cookies").exists()

        self._context = self._playwright.chromium.launch_persistent_context(
            user_data_dir=str(self.profile_dir),
            headless=False,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
            viewport={"width": 1280, "height": 800},
        )
        self._page = self._context.pages[0] if self._context.pages else self._context.new_page()
        self._page.set_default_timeout(30000)

        print("Abrindo WhatsApp Web...")
        if primeiro_uso:
            print("Primeiro uso — escaneie o QR code (ate 90s)...")
        else:
            print("Perfil salvo — entrando automaticamente (sem QR)...")

        self._page.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=30000)

        try:
            self._page.wait_for_selector(
                '#pane-side, div[aria-label="Lista de conversas"], div[aria-label="Chat list"]',
                timeout=90000,
            )
            print("WhatsApp Web pronto!\n")
        except Exception:
            print("Aviso: Nao foi possivel confirmar o login — verifique o navegador.")

    def verificar_numero(self, telefone: str) -> dict:
        """
        Verifica se o numero tem WhatsApp e detecta mensagem automatica.
        Returns: {existe: bool, bot: bool, bot_msg: str, erro: str}
        """
        page = self._page
        numero = formatar_numero_wa(telefone)
        url = (
            f"https://web.whatsapp.com/send?phone={numero}"
            "&text=&type=phone_number&app_absent=0"
        )

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            time.sleep(2)
        except Exception as e:
            return {"existe": False, "bot": False, "bot_msg": "", "erro": f"nav: {str(e)[:60]}"}

        popup_sels = [
            'div[data-testid="popup-controls"]',
            'div[data-animate-modal-body]',
            'div[role="dialog"]',
        ]
        for sel in popup_sels:
            try:
                popup = page.locator(sel).first
                if popup.is_visible(timeout=3500):
                    texto = popup.inner_text(timeout=2000).lower()
                    invalido = any(w in texto for w in [
                        "invalid", "invalido", "not found", "nao encontrado",
                        "phone number shared", "numero de telefone",
                    ])
                    if invalido:
                        for btn_sel in [
                            'button[data-testid="popup-controls-ok"]',
                            'button:has-text("OK")',
                            'button:has-text("Ok")',
                        ]:
                            try:
                                page.locator(btn_sel).first.click(timeout=2000)
                                break
                            except Exception:
                                pass
                        return {"existe": False, "bot": False, "bot_msg": "", "erro": ""}
            except Exception:
                pass

        chat_sels = [
            'div[data-testid="compose-btn-container"]',
            'footer[data-testid="conversation-footer"]',
            'div[contenteditable="true"][data-tab]',
            'div[aria-label="Digite uma mensagem"]',
            'div[aria-label="Type a message"]',
        ]
        chat_carregou = False
        for sel in chat_sels:
            try:
                page.wait_for_selector(sel, timeout=7000)
                chat_carregou = True
                break
            except Exception:
                pass

        if not chat_carregou:
            try:
                pg_text = page.inner_text("body", timeout=3000).lower()
                if any(w in pg_text for w in ["invalid", "invalido", "not a valid"]):
                    return {"existe": False, "bot": False, "bot_msg": "", "erro": ""}
            except Exception:
                pass
            return {"existe": False, "bot": False, "bot_msg": "", "erro": "timeout_chat"}

        if self.esperar_bot > 0:
            print(f"bot({self.esperar_bot}s)... ", end="", flush=True)
            time.sleep(self.esperar_bot)

        bot_detected = False
        bot_msg = ""

        incoming_sels = [
            'div.message-in span.selectable-text',
            'div[class*="message-in"] span[class*="selectable-text"]',
            'div[data-testid="msg-container"] div[class*="message-in"] span',
        ]
        for sel in incoming_sels:
            try:
                msgs = page.locator(sel).all()
                for el in msgs[-10:]:
                    try:
                        txt = el.inner_text(timeout=800).strip()
                        if txt and _eh_bot(txt):
                            bot_detected = True
                            bot_msg = txt[:300]
                            break
                    except Exception:
                        pass
                if bot_detected:
                    break
            except Exception:
                pass

        if not bot_detected:
            try:
                area = page.locator(
                    '#main div[role="application"], div[data-testid="conversation-panel-messages"]'
                ).first
                texto_area = area.inner_text(timeout=3000)
                for linha in texto_area.split("\n"):
                    linha = linha.strip()
                    if len(linha) > 15 and _eh_bot(linha):
                        bot_detected = True
                        bot_msg = linha[:300]
                        break
            except Exception:
                pass

        return {
            "existe": True,
            "bot": bot_detected,
            "bot_msg": bot_msg,
            "erro": "",
        }

    def fechar(self):
        try:
            self._context.close()
        except Exception:
            pass
        try:
            self._playwright.__exit__(None, None, None)
        except Exception:
            pass


# Processamento de planilha (verificacao de numeros)
_COL_TIPO_TEL  = "Tipo Tel"
_COL_WA_STATUS = "WA Status"

_CORES_TIPO = {
    "fixo":         "FFE4B5",
    "celular":      "E8F5E9",
    "desconhecido": "F0F0F0",
}
_CORES_WA = {
    "ativo":          "D4EDDA",
    "sem_whatsapp":   "F8D7DA",
    "bot":            "FFF3CD",
    "nao_verificado": "F0F0F0",
    "inacessivel":    "E2E3E5",
}


def processar_planilha(
    caminho: str,
    verificar_wa: bool = False,
    fixo_acao: str = "marcar",
    bot_acao: str = "ignorar",
    esperar_bot: int = 10,
    verifier: "WhatsAppVerificador | None" = None,
    pular_ja_verificados: bool = True,
) -> dict:
    """
    Processa uma planilha de leads:
      - Classifica numeros (fixo/celular)
      - Opcionalmente verifica WhatsApp + detecta bots
      - Atualiza planilha com novas colunas e acoes
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("Instale: pip install openpyxl")
        return {}

    if not Path(caminho).exists():
        print(f"Arquivo nao encontrado: {caminho}")
        return {}

    wb = openpyxl.load_workbook(caminho)

    if "Leads" not in wb.sheetnames:
        print(f"Planilha 'Leads' nao encontrada em: {caminho}")
        wb.close()
        return {}

    ws = wb["Leads"]
    hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    _HDR_FILL = PatternFill("solid", start_color="1F4E79")
    _HDR_FONT = Font(bold=True, color="FFFFFF", size=11)

    def _garantir_colunas(ws_) -> tuple:
        headers = {ws_.cell(1, c).value: c for c in range(1, ws_.max_column + 1)}
        def _add(nome):
            if nome not in headers:
                idx = ws_.max_column + 1
                c = ws_.cell(1, idx, value=nome)
                c.fill = _HDR_FILL
                c.font = _HDR_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
                headers[nome] = idx
            return headers[nome]
        return _add(_COL_TIPO_TEL), _add(_COL_WA_STATUS)

    col_tipo, col_wa = _garantir_colunas(ws)

    col_tel    = hdrs.get("Telefone")
    col_email  = hdrs.get("Email")
    col_status = hdrs.get("Status")

    if not col_tel:
        print(f"Coluna 'Telefone' nao encontrada em: {caminho}")
        wb.close()
        return {}

    stats = {
        "total": 0, "pulado": 0,
        "fixo": 0, "celular": 0, "desconhecido": 0,
        "wa_ativo": 0, "sem_whatsapp": 0, "bot": 0, "inacessivel": 0,
        "email_agendado": 0,
    }

    total_linhas = ws.max_row - 1

    for row in range(2, ws.max_row + 1):
        nome = ws.cell(row, 1).value
        if not nome:
            continue

        telefone = str(ws.cell(row, col_tel).value or "").strip()
        email    = str(ws.cell(row, col_email).value or "").strip() if col_email else ""

        stats["total"] += 1

        wa_existente = ws.cell(row, col_wa).value
        if pular_ja_verificados and wa_existente and wa_existente != "nao_verificado":
            stats["pulado"] += 1
            continue

        tipo = classificar_numero(telefone)
        stats[tipo] += 1

        cell_tipo = ws.cell(row, col_tipo)
        cell_tipo.value = tipo
        cell_tipo.fill = PatternFill("solid", start_color=_CORES_TIPO.get(tipo, "F0F0F0"))

        wa_status = "nao_verificado"

        if verificar_wa and verifier:
            if tipo == "fixo":
                wa_status = "sem_whatsapp"
                stats["sem_whatsapp"] += 1
                print(f"  [{row-1:03d}/{total_linhas}] {str(nome)[:42]:<42} fixo -> sem_whatsapp")
            elif telefone:
                print(f"  [{row-1:03d}/{total_linhas}] {str(nome)[:42]:<42} ", end="", flush=True)
                try:
                    res = verifier.verificar_numero(telefone)
                except Exception as e:
                    res = {"existe": False, "bot": False, "bot_msg": "", "erro": str(e)[:60]}

                if res.get("erro") and not res.get("existe"):
                    wa_status = "inacessivel"
                    stats["inacessivel"] += 1
                    print(f"inacessivel ({res['erro'][:40]})")
                elif not res.get("existe"):
                    wa_status = "sem_whatsapp"
                    stats["sem_whatsapp"] += 1
                    print("sem WhatsApp")
                elif res.get("bot"):
                    wa_status = "bot"
                    stats["bot"] += 1
                    trecho = res["bot_msg"][:60].replace("\n", " ")
                    print(f"bot: {trecho}")
                else:
                    wa_status = "ativo"
                    stats["wa_ativo"] += 1
                    print("WA ativo")
            else:
                wa_status = "sem_whatsapp"
        elif tipo == "fixo":
            wa_status = "sem_whatsapp"

        cell_wa = ws.cell(row, col_wa)
        cell_wa.value = wa_status
        cell_wa.fill = PatternFill("solid", start_color=_CORES_WA.get(wa_status, "F0F0F0"))

        if tipo == "fixo":
            if fixo_acao == "email":
                if col_status:
                    status_atual = ws.cell(row, col_status).value
                    if status_atual in ("", None):
                        ws.cell(row, col_status).value = "novo"
                if email:
                    stats["email_agendado"] += 1

        if wa_status == "bot" and col_status:
            status_atual = ws.cell(row, col_status).value
            if bot_acao == "ignorar":
                if status_atual not in ("email_enviado", "respondeu", "fechado"):
                    ws.cell(row, col_status).value = "bot_ignorado"
                    ws.cell(row, col_status).fill = PatternFill("solid", start_color="E2E3E5")
            elif bot_acao == "email":
                if status_atual in ("", None, "bot_ignorado"):
                    ws.cell(row, col_status).value = "novo"
                if email:
                    stats["email_agendado"] += 1

        if stats["total"] % 10 == 0:
            wb.save(caminho)

    wb.save(caminho)
    return stats


# WhatsApp Web — Auto Conversador
class AutoConversador:
    """
    Monitora WhatsApp Web e responde mensagens de leads autorizados via Ollama.
    """

    _SEL_CHAT_LIST = [
        '#pane-side',
        'div[aria-label="Lista de conversas"]',
        'div[aria-label="Chat list"]',
    ]
    _SEL_UNREAD_BADGE = [
        'span[data-testid="icon-unread-count"]',
        'span[aria-label*="mensagem nao lida"]',
        'span[aria-label*="unread message"]',
        'div[aria-label*="nao lida"]',
    ]
    _SEL_CHAT_INPUT = [
        'div[data-testid="conversation-compose-box-input"]',
        'div[contenteditable="true"][data-tab]',
        'div[aria-label="Digite uma mensagem"]',
        'div[aria-label="Type a message"]',
        'footer div[contenteditable="true"]',
    ]
    _SEL_CHAT_LOADED = [
        'div[data-testid="compose-btn-container"]',
        'footer[data-testid="conversation-footer"]',
        'div[aria-label="Digite uma mensagem"]',
        'div[aria-label="Type a message"]',
    ]

    def __init__(
        self,
        modelo: str = MODELO_PADRAO,
        ollama_url: str = OLLAMA_URL,
        intervalo: int = 30,
        pausa_resposta: int = 5,
        profile_dir: str = WA_PROFILE_DIR,
    ):
        self.modelo = modelo
        self.ollama_url = ollama_url
        self.intervalo = intervalo
        self.pausa_resposta = pausa_resposta
        self.profile_dir = Path(profile_dir)

        self._playwright = None
        self._context = None
        self._page = None

        self.autorizados: dict = {}
        self.historico: dict = {}
        self.agente_cfg: dict = carregar_agente_cfg()
        self._rodando = True
        self._processados_ciclo: set = set()

    def _iniciar_whatsapp(self):
        from playwright.sync_api import sync_playwright

        self._playwright = sync_playwright().__enter__()
        self.profile_dir.mkdir(parents=True, exist_ok=True)

        primeiro_uso = not (self.profile_dir / "Default" / "Cookies").exists()

        self._context = self._playwright.chromium.launch_persistent_context(
            user_data_dir=str(self.profile_dir),
            headless=False,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
            viewport={"width": 1280, "height": 800},
        )
        self._page = self._context.pages[0] if self._context.pages else self._context.new_page()
        self._page.set_default_timeout(30000)

        print("Abrindo WhatsApp Web...")
        if primeiro_uso:
            print("Primeiro uso — escaneie o QR code (ate 90s)...")
        else:
            print("Perfil salvo — entrando automaticamente...")

        self._page.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=30000)

        try:
            self._page.wait_for_selector(
                ", ".join(self._SEL_CHAT_LIST),
                timeout=90000,
            )
            print("WhatsApp Web pronto!\n")
        except Exception:
            print("Aviso: Nao confirmou o login — verifique o navegador e pressione Enter.")
            input()

    def fechar(self):
        try:
            self._context.close()
        except Exception:
            pass
        try:
            self._playwright.__exit__(None, None, None)
        except Exception:
            pass

    def _ler_mensagens_chat(self) -> list:
        """Lê as mensagens visiveis do chat aberto."""
        page = self._page
        mensagens = []

        sels_container = [
            ('div.message-in, div.message-out', 'message-in'),
            ('div[class*="message-in"], div[class*="message-out"]', 'message-in'),
        ]

        for sel, cls_in in sels_container:
            try:
                containers = page.locator(sel).all()
                if not containers:
                    continue

                for container in containers[-25:]:
                    try:
                        classe = container.get_attribute("class", timeout=500) or ""
                        de = "eles" if cls_in in classe else "nos"

                        texto_el = container.locator('span.selectable-text').first
                        texto = texto_el.inner_text(timeout=800).strip()

                        if texto:
                            mensagens.append({"de": de, "texto": texto})
                    except Exception:
                        pass

                if mensagens:
                    return mensagens
            except Exception:
                pass

        try:
            in_sels = [
                'div[class*="message-in"] span[class*="selectable-text"]',
                'div.message-in span.selectable-text',
            ]
            out_sels = [
                'div[class*="message-out"] span[class*="selectable-text"]',
                'div.message-out span.selectable-text',
            ]

            msgs_in, msgs_out = [], []

            for sel in in_sels:
                try:
                    els = page.locator(sel).all()
                    msgs_in = [e.inner_text(timeout=500).strip() for e in els[-15:] if e]
                    if msgs_in:
                        break
                except Exception:
                    pass

            for sel in out_sels:
                try:
                    els = page.locator(sel).all()
                    msgs_out = [e.inner_text(timeout=500).strip() for e in els[-15:] if e]
                    if msgs_out:
                        break
                except Exception:
                    pass

            for t in msgs_in:
                if t:
                    mensagens.append({"de": "eles", "texto": t})
            for t in msgs_out:
                if t:
                    mensagens.append({"de": "nos", "texto": t})

        except Exception:
            pass

        return mensagens

    _JS_EXTRAIR_NUMERO = """
        () => {
            const RE = /(\\d{10,15})@c\\.us/;
            const main = document.querySelector('#main');
            if (main) {
                const rows = main.querySelectorAll('[data-id]');
                for (const r of rows) {
                    const m = (r.getAttribute('data-id') || '').match(RE);
                    if (m) return m[1];
                }
            }
            for (const el of document.querySelectorAll('[data-id*="@c.us"]')) {
                const m = (el.getAttribute('data-id') || '').match(RE);
                if (m) return m[1];
            }
            for (const el of document.querySelectorAll('[data-jid]')) {
                const m = (el.getAttribute('data-jid') || '').match(RE);
                if (m) return m[1];
            }
            const hdr = document.querySelector('#main header');
            if (hdr) {
                for (const el of hdr.querySelectorAll('[aria-label]')) {
                    const clean = (el.getAttribute('aria-label') || '').replace(/\\D/g, '');
                    if (clean.length >= 10 && clean.length <= 15) return clean;
                }
            }
            const um = window.location.href.match(/phone=(\\d{10,15})/);
            if (um) return um[1];
            return null;
        }
    """

    def _extrair_numero_do_chat(self, debug: bool = False) -> str:
        page = self._page

        for sel in [
            '#main div[data-id]',
            '#main [data-id*="@c.us"]',
            '#main [data-jid]',
            '#main footer',
        ]:
            try:
                page.wait_for_selector(sel, timeout=5000)
                break
            except Exception:
                pass

        try:
            result = page.evaluate(self._JS_EXTRAIR_NUMERO)
            if result:
                nums = re.sub(r"\D", "", str(result))
                if 10 <= len(nums) <= 15:
                    if not nums.startswith("55"):
                        nums = "55" + nums
                    return nums
        except Exception:
            pass

        try:
            url = page.url or ""
            m = re.search(r"phone=(\d{10,15})", url)
            if m:
                nums = m.group(1)
                if not nums.startswith("55"):
                    nums = "55" + nums
                return nums
        except Exception:
            pass

        return ""

    def _enviar_mensagem_chat(self, texto: str) -> bool:
        """Envia uma mensagem no chat atualmente aberto."""
        page = self._page

        input_el = None
        for sel in self._SEL_CHAT_INPUT:
            try:
                el = page.locator(sel).first
                if el.is_visible(timeout=2000):
                    input_el = el
                    break
            except Exception:
                pass

        if not input_el:
            print("  Campo de texto nao encontrado")
            return False

        try:
            input_el.click()
            time.sleep(0.4)

            linhas = texto.split("\n")
            for i, linha in enumerate(linhas):
                if i > 0:
                    input_el.press("Shift+Enter")
                    time.sleep(0.1)
                if linha:
                    input_el.type(linha, delay=15)

            time.sleep(0.6)
            input_el.press("Enter")
            time.sleep(1.2)
            return True

        except Exception as e:
            print(f"  Erro ao enviar mensagem: {e}")
            return False

    def _aguardar_chat_carregar(self) -> bool:
        for sel in self._SEL_CHAT_LOADED:
            try:
                self._page.wait_for_selector(sel, timeout=8000)
                return True
            except Exception:
                pass
        return False

    @staticmethod
    def _normalizar_numero(numero: str) -> list:
        n = re.sub(r"\D", "", numero)
        if not n.startswith("55"):
            n = "55" + n
        variacoes = [n]
        if len(n) == 12:
            variacoes.append(n[:4] + "9" + n[4:])
        elif len(n) == 13 and n[4] == "9":
            variacoes.append(n[:4] + n[5:])
        return list(dict.fromkeys(variacoes))

    def _navegar_para_numero(self, numero: str) -> bool:
        for num_var in self._normalizar_numero(numero):
            url = (
                f"https://web.whatsapp.com/send?phone={num_var}"
                "&text=&type=phone_number&app_absent=0"
            )
            try:
                self._page.goto(url, wait_until="domcontentloaded", timeout=25000)
                time.sleep(2)
            except Exception:
                continue

            try:
                popup = self._page.locator(
                    'div[data-animate-modal-body], div[role="dialog"]'
                ).first
                if popup.is_visible(timeout=2000):
                    txt = popup.inner_text(timeout=1000).lower()
                    if any(w in txt for w in ["invalido", "invalid", "nao esta", "not registered"]):
                        try:
                            self._page.locator('button:has-text("OK"), button:has-text("Ok")').first.click(timeout=1500)
                        except Exception:
                            pass
                        continue
            except Exception:
                pass

            if self._aguardar_chat_carregar():
                return True

        return False

    def _voltar_para_lista(self):
        try:
            self._page.goto(
                "https://web.whatsapp.com",
                wait_until="domcontentloaded",
                timeout=15000,
            )
            time.sleep(1)
        except Exception:
            pass

    def _processar_chat(self, numero: str, info: dict) -> bool:
        mensagens = self._ler_mensagens_chat()
        if not mensagens:
            return False

        hist = self.historico.get(numero, {"mensagens": [], "status": "ativo"})

        if hist.get("status") in ("sem_interesse", "fechado"):
            return False

        msgs_deles = [m for m in mensagens if m["de"] == "eles"]
        if not msgs_deles:
            return False

        ultima = msgs_deles[-1]
        msg_hash = _hash_msg(ultima["texto"])

        if msg_hash == hist.get("ultimo_hash_recebido"):
            return False

        if _eh_bot(ultima["texto"]):
            print(f"  Mensagem automatica detectada — registrando sem responder")
            hist.setdefault("mensagens", []).append({
                "de":    "eles",
                "texto": ultima["texto"],
                "ts":    datetime.now().strftime("%Y-%m-%d %H:%M"),
                "tipo":  "bot",
            })
            hist["ultimo_hash_recebido"] = msg_hash
            self.historico[numero] = hist
            salvar_historico(self.historico)
            return False

        print(f"  Mensagem: {ultima['texto'][:90]}")

        if _eh_sem_interesse(ultima["texto"]):
            print(f"  Lead sinalizou sem interesse — encerrando cortesmente")
            msg_encerramento = gerar_resposta_ollama(
                ollama_url=self.ollama_url,
                modelo=self.modelo,
                numero=numero,
                info=info,
                historico=self.historico,
                nova_mensagem=ultima["texto"],
                agente_cfg=self.agente_cfg,
            )
            if msg_encerramento:
                time.sleep(self.pausa_resposta)
                self._enviar_mensagem_chat(msg_encerramento)

            hist.setdefault("mensagens", []).extend([
                {"de": "eles", "texto": ultima["texto"],
                 "ts": datetime.now().strftime("%Y-%m-%d %H:%M"), "tipo": "sem_interesse"},
                {"de": "nos", "texto": msg_encerramento,
                 "ts": datetime.now().strftime("%Y-%m-%d %H:%M")},
            ])
            hist["ultimo_hash_recebido"] = msg_hash
            hist["status"] = "sem_interesse"
            self.historico[numero] = hist
            salvar_historico(self.historico)
            return True

        print(f"  Gerando resposta ({self.modelo})...")
        resposta = gerar_resposta_ollama(
            ollama_url=self.ollama_url,
            modelo=self.modelo,
            numero=numero,
            info=info,
            historico=self.historico,
            nova_mensagem=ultima["texto"],
            agente_cfg=self.agente_cfg,
        )

        if not resposta:
            print(f"  Nao gerou resposta — pulando")
            return False

        print(f"  Resposta: {resposta[:100]}{'...' if len(resposta) > 100 else ''}")

        time.sleep(self.pausa_resposta)

        if not self._enviar_mensagem_chat(resposta):
            print(f"  Falha ao enviar")
            return False

        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        hist.setdefault("mensagens", []).extend([
            {"de": "eles", "texto": ultima["texto"], "ts": ts, "tipo": "humano"},
            {"de": "nos",  "texto": resposta,         "ts": ts},
        ])
        hist["ultimo_hash_recebido"] = msg_hash
        hist["ultima_resposta_ts"]   = ts
        self.historico[numero] = hist
        salvar_historico(self.historico)

        print(f"  Resposta enviada!")
        return True

    _JS_UNREAD_NUMBERS = """
        () => {
            const BADGE_SELS = [
                '[data-testid="icon-unread-count"]',
                'span[aria-label*="mensagem nao lida"]',
                'span[aria-label*="unread message"]',
                'div[aria-label*="nao lida"]',
            ];
            const RE_NUM   = /(\\d{10,15})@c\\.us/;
            const RE_PLAIN = /^[+\\d][\\d\\s\\-().]{7,14}$/;
            function fromDataId(el) {
                const di = el.getAttribute('data-id') || '';
                const m  = di.match(RE_NUM);
                return m ? m[1] : null;
            }
            function fromTitle(container) {
                const spans = container.querySelectorAll('span[title], [title]');
                for (const s of spans) {
                    const t = (s.getAttribute('title') || '').trim();
                    const clean = t.replace(/\\D/g, '');
                    if (clean.length >= 10 && clean.length <= 15 && RE_PLAIN.test(t)) {
                        return clean;
                    }
                }
                return null;
            }
            const seen = new Set();
            const results = [];
            for (const sel of BADGE_SELS) {
                for (const badge of document.querySelectorAll(sel)) {
                    let num = null;
                    let el = badge;
                    for (let i = 0; i < 10; i++) {
                        el = el.parentElement;
                        if (!el) break;
                        num = fromDataId(el);
                        if (num) break;
                        const child = el.querySelector('[data-id*="@c.us"]');
                        if (child) { num = fromDataId(child); if (num) break; }
                        num = fromTitle(el);
                        if (num) break;
                    }
                    if (num && !seen.has(num)) {
                        seen.add(num);
                        results.push(num);
                    }
                }
                if (results.length) break;
            }
            return results;
        }
    """

    def _monitorar_sidebar(self) -> int:
        page = self._page
        respostas = 0
        numeros_raw: list = []
        try:
            numeros_raw = page.evaluate(self._JS_UNREAD_NUMBERS) or []
        except Exception as e:
            print(f"  JS sidebar falhou: {e}")

        if numeros_raw:
            print(f"  {len(numeros_raw)} numero(s) com mensagem nao lida detectado(s)")
        else:
            total_badges = 0
            for sel in self._SEL_UNREAD_BADGE:
                try:
                    n = page.locator(sel).count()
                    if n:
                        total_badges = n
                        break
                except Exception:
                    pass

        for num_raw in numeros_raw:
            variacoes = self._normalizar_numero(str(num_raw))
            numero = None
            info   = None
            for v in variacoes:
                info = self.autorizados.get(v)
                if info:
                    numero = v
                    break

            if not numero or not info:
                print(f"  {variacoes} — nao autorizado")
                continue

            if numero in self._processados_ciclo:
                continue

            nome = info.get("nome") or numero
            print(f"  Autorizado: {nome} ({numero})")

            if not self._navegar_para_numero(numero):
                print(f"  Nao conseguiu abrir chat de {numero}")
                continue

            if self._processar_chat(numero, info):
                respostas += 1
                self._processados_ciclo.add(numero)

            self._voltar_para_lista()
            time.sleep(1)

        return respostas

    def _verificar_contatos_ativos(self):
        ativos = {
            num: info
            for num, info in self.autorizados.items()
            if (
                num in self.historico
                and self.historico[num].get("status") not in ("sem_interesse", "fechado")
                and any(m["de"] == "eles" for m in self.historico[num].get("mensagens", []))
                and num not in self._processados_ciclo
            )
        }

        if not ativos:
            return

        print(f"\n  Verificando {len(ativos)} contato(s) com conversa ativa...")

        for numero, info in ativos.items():
            nome = info.get("nome") or numero
            print(f"  -> {nome[:45]}...", end="", flush=True)

            try:
                if not self._navegar_para_numero(numero):
                    print(" inacessivel")
                    continue

                respondeu = self._processar_chat(numero, info)
                print(" " + ("respondido" if respondeu else "sem novidades"))

                if respondeu:
                    self._processados_ciclo.add(numero)

            except Exception as e:
                print(f" {str(e)[:50]}")

            self._voltar_para_lista()
            time.sleep(2)

    def rodar(self):
        """Loop principal de monitoramento do WhatsApp."""
        print(f"\n{'='*62}")
        print(f"  AUTO CONVERSA — WhatsApp + Ollama ({self.modelo})")
        print(f"{'='*62}")
        print(f"  Contatos autorizados : {len(self.autorizados)}")
        print(f"  Conversas no historico: {len(self.historico)}")
        print(f"  Intervalo de varredura: {self.intervalo}s")
        print(f"  Pausa antes de responder: {self.pausa_resposta}s")
        print(f"  [Ctrl+C para parar com seguranca]\n")

        ciclo = 0

        while self._rodando:
            ciclo += 1
            self._processados_ciclo = set()
            agora = datetime.now().strftime("%H:%M:%S")

            print(f"\n[{agora}] -- Ciclo #{ciclo} --")

            try:
                n_respostas = self._monitorar_sidebar()

                if n_respostas == 0:
                    print("  Nenhuma mensagem nova no sidebar")

                if ciclo % 5 == 0:
                    self._verificar_contatos_ativos()

            except Exception as e:
                print(f"  Aviso: Erro no ciclo #{ciclo}: {e}")

            print(f"  Proxima varredura em {self.intervalo}s...")
            for _ in range(self.intervalo):
                if not self._rodando:
                    break
                time.sleep(1)

        print("\nEncerrando auto conversa...")
        salvar_historico(self.historico)
        print("Historico salvo. Ate mais!")


# CLI
def main():
    parser = argparse.ArgumentParser(
        description="Verificacao WhatsApp + Auto Conversa com Ollama",
    )

    subparsers = parser.add_subparsers(dest="comando")

    # Sub-comando: verificar
    p_ver = subparsers.add_parser("verificar", help="Verifica numeros nas planilhas")
    p_ver.add_argument("--leads", "-l", default=None)
    p_ver.add_argument("--pasta", "-p", default=None)
    p_ver.add_argument("--verificar-wa", "-v", action="store_true")
    p_ver.add_argument("--so-classificar", "-s", action="store_true")
    p_ver.add_argument("--fixo-acao", default="marcar", choices=["marcar", "email"])
    p_ver.add_argument("--bot-acao", default="ignorar", choices=["ignorar", "email"])
    p_ver.add_argument("--esperar-bot", type=int, default=10)

    # Sub-comando: conversa
    p_conv = subparsers.add_parser("conversa", help="Auto Conversa com Ollama")
    p_conv.add_argument("--modelo", "-m", default=MODELO_PADRAO)
    p_conv.add_argument("--ollama-url", default=OLLAMA_URL)
    p_conv.add_argument("--intervalo", "-i", type=int, default=30)
    p_conv.add_argument("--pausa", "-p", type=int, default=5)

    args = parser.parse_args()

    if args.comando == "conversa":
        autorizados = carregar_autorizados()
        if not autorizados:
            print(f"\nNenhum numero autorizado em '{LOG_ENVIADOS}'.")
            sys.exit(1)

        print(f"\n{'='*62}")
        print(f"  AUTO CONVERSA — WhatsApp + Ollama")
        print(f"{'='*62}")
        print(f"  Verificando Ollama ({args.modelo})...")

        if not verificar_ollama(args.ollama_url, args.modelo):
            print(f"\nOllama nao disponivel em {args.ollama_url}")
            sys.exit(1)

        agente = AutoConversador(
            modelo=args.modelo,
            ollama_url=args.ollama_url,
            intervalo=args.intervalo,
            pausa_resposta=args.pausa,
        )
        agente.autorizados = autorizados
        agente.historico   = carregar_historico()

        def _sair(sig, frame):
            print("\n\nSinal de parada recebido — encerrando no proximo ciclo...")
            agente._rodando = False

        signal.signal(signal.SIGINT, _sair)

        try:
            agente._iniciar_whatsapp()
            agente.rodar()
        finally:
            agente.fechar()
            salvar_historico(agente.historico)
            print("Historico salvo.")

    elif args.comando == "verificar":
        if args.pasta:
            pasta = Path(args.pasta)
            arquivos = sorted(pasta.glob("*.xlsx"))
        elif args.leads:
            p = Path(args.leads)
            arquivos = [p] if p.exists() else sorted(Path(".").glob(args.leads))
        else:
            parser.print_help()
            sys.exit(1)

        if not arquivos:
            print("Nenhum arquivo .xlsx encontrado.")
            sys.exit(1)

        verifier = None
        if args.verificar_wa and not args.so_classificar:
            verifier = WhatsAppVerificador(esperar_bot=args.esperar_bot)
            try:
                verifier.iniciar()
            except Exception as e:
                print(f"Erro ao iniciar WhatsApp Web: {e}")
                sys.exit(1)

        try:
            for arquivo in arquivos:
                print(f"{arquivo.name}")
                stats = processar_planilha(
                    caminho=str(arquivo),
                    verificar_wa=(args.verificar_wa and not args.so_classificar),
                    fixo_acao=args.fixo_acao,
                    bot_acao=args.bot_acao,
                    esperar_bot=args.esperar_bot,
                    verifier=verifier,
                )
                print(f"  Total: {stats.get('total',0)} | Fixo: {stats.get('fixo',0)} | Celular: {stats.get('celular',0)}")
        finally:
            if verifier:
                verifier.fechar()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
