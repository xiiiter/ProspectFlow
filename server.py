#!/usr/bin/env python3
"""
Servidor Flask do painel de prospeccao.

Rode com:
    python server.py
"""

from flask import Flask, render_template, jsonify, request, Response, send_file
import threading
import queue
import json
import sys
import re
import time
import os
import socket
import random
from pathlib import Path
from datetime import datetime, date
import importlib.util
import webbrowser

import importlib.util as ilu

try:
    from core.enricher import (
        enriquecer_cnpj, buscar_dono_telefone, extrair_emails_do_site,
        buscar_redes_sociais, analisar_website, analisar_seo,
        calcular_score_oportunidade, classificar_telefone
    )
    from core.proposal import gerar_proposta_html, PRECOS_PADRAO
    ENRICHER_OK = True
except Exception as _e:
    ENRICHER_OK = False
    print(f"[warn] core/enricher nao carregado: {_e}")


def _get_local_ip() -> str:
    """Retorna o IP local da maquina na rede Wi-Fi/LAN."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

BASE_DIR  = Path(__file__).parent
LEADS_DIR = BASE_DIR / "data" / "leads"
LOGS_DIR  = BASE_DIR / "data" / "logs"
TMPL_DIR  = BASE_DIR / "templates"
TMPL_DIR.mkdir(exist_ok=True)
LEADS_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)

_CRM_FILE = LOGS_DIR / "crm_status.json"

app = Flask(__name__, template_folder=str(TMPL_DIR))
app.config["JSON_AS_ASCII"] = False

# Estado global
_state = {
    "running":      False,
    "should_stop":  False,
    "logs":         [],
    "stats": {"total":0,"sem_site":0,"social":0,"site":0,
              "wpp":0,"cnpj":0,"email":0,"insta":0},
    "progress": {"current":0,"total":1,"percent":0,"label":""},
    "captcha_pendente": False,
    "captcha_evento": threading.Event(),
    "captcha_msg": "",
}
_state["captcha_evento"].set()
_subscribers: list[queue.Queue] = []

# Estado de envio automatico
_send_state = {
    "running":     False,
    "should_stop": False,
    "logs":        [],
    "stats": {"total": 0, "emails": 0, "whatsapp": 0, "errors": 0},
    "progress": {"current": 0, "total": 1, "percent": 0, "label": ""},
    # Rate limiting — atualizado em tempo real pelo worker
    "rate": {
        "limite_diario":  50,
        "enviados_hoje":  0,
        "delay_min":      20,
        "delay_max":      45,
        "pausa_a_cada":   10,
        "pausa_duracao":  300,
        "pausando_ate":   "",
    },
}
_send_subscribers: list[queue.Queue] = []

# Estado Auto Conversa
_autoconv_state = {
    "running":  False,
    "should_stop": False,
    "logs":     [],
    "modelo":   "llama3",
    "intervalo": 30,
    "pausa":    5,
    "stats": {"respondidos": 0, "bots_ignorados": 0, "sem_interesse": 0, "erros": 0},
}
_autoconv_subscribers: list[queue.Queue] = []
_autoconv_agente = None   # referencia ao AutoConversador em execucao

# Estado CRM
_crm_state = {}  # numero/email -> {status, notas, ultima_atualizacao}

PROJECT_URLS  = {"xiiter": "https://xiiter.vercel.app/", "inserth": "https://www.inserth.com"}
PROJECT_NAMES = {"xiiter": "Xiiter", "inserth": "Inserth"}

# Carrega modulos do projeto
def _load_mod(name: str, path: Path):
    try:
        spec = importlib.util.spec_from_file_location(name, path)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod
    except Exception as e:
        print(f"[warn] Nao carregou {path.name}: {e}")
        return None


try:
    from core.discovery import NICHOS_BRASIL as NICHOS_BR, CIDADES_BRASIL as CIDADES_BR
    from core.discovery import NICHOS_GRINGA as NICHOS_US, CIDADES_USA as CIDADES_US
except Exception as _e:
    print(f"[warn] core/discovery nao carregado: {_e}")
    NICHOS_BR  = {}
    CIDADES_BR = {}
    NICHOS_US  = {}
    CIDADES_US = {}

PRECOS_DEFAULT = {
    "dentista":2500,"clinica odontologica":3000,"ortodontista":2800,
    "fisioterapeuta":1800,"psicologo":2000,"nutricionista":1800,
    "medico":3000,"clinica medica":3000,"dermatologista":2800,
    "oftalmologista":2500,"pediatra":2800,"veterinario":2200,
    "clinica veterinaria":2500,"pet shop":1500,
    "barbearia":1200,"salao de beleza":1500,"cabeleireiro":1400,
    "manicure":900,"spa":2000,"estetica":1800,
    "academia":2000,"crossfit":2000,"pilates":1800,"yoga":1500,
    "restaurante":1800,"lanchonete":1200,"pizzaria":1500,
    "padaria":1300,"cafeteria":1200,"hamburgueria":1300,
    "advogado":3500,"contador":2800,"imobiliaria":3000,
    "arquiteto":3000,"escola":2500,"cursinho":2000,
    "construtora":3000,"eletricista":1500,"mecanica":1800,
    "dentist":800,"barbershop":500,"restaurant":600,
    "gym":600,"salon":500,"law":1200,"accountant":800,
    "default":2000,
}


def _notify(msg: dict):
    dead = []
    for q in _subscribers:
        try:
            q.put_nowait(msg)
        except Exception:
            dead.append(q)
    for q in dead:
        try:
            _subscribers.remove(q)
        except ValueError:
            pass


def _log(msg: str, level: str = "info"):
    entry = {"type":"log","msg":msg,"level":level,"ts":datetime.now().strftime("%H:%M:%S")}
    _state["logs"].append(entry)
    if len(_state["logs"]) > 3000:
        _state["logs"] = _state["logs"][-1500:]
    _notify(entry)


def _push_stats():
    _notify({"type":"stats","data":dict(_state["stats"])})


def _push_progress():
    _notify({"type":"progress","data":dict(_state["progress"])})


def _send_notify(msg: dict):
    dead = []
    for q in _send_subscribers:
        try:
            q.put_nowait(msg)
        except Exception:
            dead.append(q)
    for q in dead:
        try:
            _send_subscribers.remove(q)
        except ValueError:
            pass


def _send_log(msg: str, level: str = "info"):
    entry = {"type":"log","msg":msg,"level":level,"ts":datetime.now().strftime("%H:%M:%S")}
    _send_state["logs"].append(entry)
    if len(_send_state["logs"]) > 2000:
        _send_state["logs"] = _send_state["logs"][-1000:]
    _send_notify(entry)


def _autoconv_notify(msg: dict):
    dead = []
    for q in _autoconv_subscribers:
        try:
            q.put_nowait(msg)
        except Exception:
            dead.append(q)
    for q in dead:
        try: _autoconv_subscribers.remove(q)
        except ValueError: pass


def _autoconv_log(msg: str, level: str = "info"):
    entry = {"type":"log","msg":msg,"level":level,"ts":datetime.now().strftime("%H:%M:%S")}
    _autoconv_state["logs"].append(entry)
    if len(_autoconv_state["logs"]) > 1000:
        _autoconv_state["logs"] = _autoconv_state["logs"][-500:]
    _autoconv_notify(entry)


# Rotas
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/data")
def get_data():
    return jsonify({
        "states_br": sorted(CIDADES_BR.keys()),
        "states_us": sorted(CIDADES_US.keys()),
        "nichos_br": NICHOS_BR,
        "nichos_us": NICHOS_US,
        "precos":    PRECOS_DEFAULT,
    })


@app.route("/api/cities")
def get_cities():
    modo   = request.args.get("modo", "brasil")
    estado = request.args.get("estado", "")
    db     = CIDADES_BR if modo == "brasil" else CIDADES_US
    return jsonify({"cities": sorted(db.get(estado, []))})


@app.route("/api/start", methods=["POST"])
def start():
    if _state["running"]:
        return jsonify({"error":"Ja esta rodando"}), 400

    data        = request.json or {}
    cidades     = data.get("cidades", [])
    nichos      = data.get("nichos",  [])
    max_leads   = int(data.get("max_leads",  30))
    auto_save   = int(data.get("auto_save",   5))
    buscar_cnpj       = bool(data.get("cnpj",     True))
    precos            = data.get("precos", {})
    gerar_email       = bool(data.get("gerar_email", False))
    moeda             = data.get("moeda", "BRL")
    retype             = bool(data.get("retype", False))
    retype_prioridades = data.get("retype_prioridades", ["Alta"])

    if not cidades or not nichos:
        return jsonify({"error":"Selecione cidades e nichos"}), 400

    try:
        with open(BASE_DIR / "config.json", "w", encoding="utf-8") as f:
            json.dump({"moeda":moeda,"precos":precos,"max_leads":max_leads,"auto_save":auto_save},
                      f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    _state["running"]     = True
    _state["should_stop"] = False
    _state["logs"]        = []
    for k in _state["stats"]:
        _state["stats"][k] = 0
    total = len(cidades) * len(nichos)
    _state["progress"] = {"current":0,"total":total,"percent":0,"label":""}

    params = {"cidades":cidades,"nichos":nichos,"max_leads":max_leads,
              "auto_save":auto_save,"cnpj":buscar_cnpj,"precos":precos,
              "gerar_email":gerar_email,"moeda":moeda,
              "retype":retype,"retype_prioridades":retype_prioridades}

    threading.Thread(target=_worker, args=(params,), daemon=True).start()
    return jsonify({"ok":True,"total":total})


@app.route("/api/stop", methods=["POST"])
def stop():
    _state["should_stop"] = True
    _notify({"type":"running","data":False})
    return jsonify({"ok":True})


@app.route("/api/stream")
def stream():
    """Server-Sent Events para atualizacoes em tempo real."""
    q = queue.Queue()
    _subscribers.append(q)

    def generate():
        yield f"data: {json.dumps({'type':'init','stats':dict(_state['stats']),'running':_state['running'],'progress':dict(_state['progress'])})}\n\n"
        for log in _state["logs"][-100:]:
            yield f"data: {json.dumps(log)}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield f"data: {json.dumps(msg)}\n\n"
                except queue.Empty:
                    yield "data: {\"type\":\"ping\"}\n\n"
        finally:
            try:
                _subscribers.remove(q)
            except ValueError:
                pass

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no",
                             "Connection":"keep-alive"})


@app.route("/api/status")
def status():
    return jsonify({"running":_state["running"],"stats":dict(_state["stats"]),
                    "progress":dict(_state["progress"])})


@app.route("/api/leads")
def list_leads():
    files = []
    for f in sorted(LEADS_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            stat  = f.stat()
            rows  = "?"
            nicho = cidade = ""
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
                ws = wb["Leads"]
                rows = max(0, ws.max_row - 1)
                nicho  = str(ws.cell(2, 2).value or "")
                cidade = str(ws.cell(2, 3).value or "")
                wb.close()
            except Exception:
                pass
            files.append({
                "name":  f.name,
                "nicho": nicho,
                "cidade": cidade,
                "size":  _fmt_size(stat.st_size),
                "date":  datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M"),
                "rows":  rows,
            })
        except Exception:
            pass
    return jsonify({"files": files})


@app.route("/api/leads/preview/<path:filename>")
def preview_lead(filename):
    safe = Path(filename).name
    path = LEADS_DIR / safe
    if not path.exists():
        return jsonify({"error":"Nao encontrado"}), 404
    try:
        import openpyxl
        wb   = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws   = wb["Leads"]
        hdrs = [ws.cell(1, c).value for c in range(1, min(ws.max_column+1, 20))]
        rows = []
        for r in range(2, min(ws.max_row+1, 52)):
            row = [str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column+1, 20))]
            if any(v for v in row):
                rows.append(row)
        wb.close()
        return jsonify({"headers":hdrs,"rows":rows,"total":ws.max_row-1})
    except Exception as e:
        return jsonify({"error":str(e)}), 500


@app.route("/api/leads/download/<path:filename>")
def download_lead(filename):
    safe = Path(filename).name
    path = LEADS_DIR / safe
    if not path.exists():
        return jsonify({"error":"Nao encontrado"}), 404
    return send_file(str(path), as_attachment=True, download_name=safe)


def _fmt_size(b: int) -> str:
    if b < 1024:    return f"{b} B"
    if b < 1048576: return f"{b/1024:.1f} KB"
    return f"{b/1048576:.1f} MB"


# Rotas de Envio Automatico
@app.route("/api/send/files")
def send_files():
    files = []
    for f in sorted(LEADS_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
        files.append({"name": f.name, "path": str(f), "dir": "leads"})
    for f in sorted(BASE_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
        files.append({"name": f.name, "path": str(f), "dir": "raiz"})
    return jsonify({"files": files})


@app.route("/api/send/start", methods=["POST"])
def send_start():
    if _send_state["running"]:
        return jsonify({"error": "Ja esta rodando"}), 400
    data      = request.json or {}
    files     = data.get("files", [])
    channels  = data.get("channels", [])
    project   = data.get("project", "xiiter")
    priorities= data.get("priorities", ["Alta", "Media"])
    max_sends = int(data.get("max_sends", 20))
    templates = data.get("templates", {})
    limite_diario  = int(data.get("limite_diario",  50))
    delay_min      = int(data.get("delay_min",       20))
    delay_max      = int(data.get("delay_max",       45))
    pausa_a_cada   = int(data.get("pausa_a_cada",    10))
    pausa_duracao  = int(data.get("pausa_duracao",  300))
    nichos_filtro  = data.get("nichos_filtro",  [])
    cidades_filtro = data.get("cidades_filtro", [])
    if not files:    return jsonify({"error": "Selecione pelo menos uma planilha"}), 400
    if not channels: return jsonify({"error": "Selecione pelo menos um canal"}), 400
    _send_state["running"]     = True
    _send_state["should_stop"] = False
    _send_state["logs"]        = []
    for k in _send_state["stats"]: _send_state["stats"][k] = 0
    _send_state["progress"] = {"current":0,"total":max_sends,"percent":0,"label":""}
    _send_state["rate"].update({
        "limite_diario": limite_diario,
        "delay_min":     delay_min,
        "delay_max":     delay_max,
        "pausa_a_cada":  pausa_a_cada,
        "pausa_duracao": pausa_duracao,
        "pausando_ate":  "",
    })
    auto_conversa    = bool(data.get("auto_conversa", False))
    ac_modelo        = data.get("ac_modelo",     "llama3")
    ac_ollama_url    = data.get("ac_ollama_url", "http://localhost:11434")
    ac_intervalo     = int(data.get("ac_intervalo", 30))
    ac_pausa         = int(data.get("ac_pausa",     5))

    params = {"files":files,"channels":channels,"project":project,
              "priorities":priorities,"max_sends":max_sends,"templates":templates,
              "limite_diario":limite_diario,"delay_min":delay_min,"delay_max":delay_max,
              "pausa_a_cada":pausa_a_cada,"pausa_duracao":pausa_duracao,
              "nichos_filtro":nichos_filtro,"cidades_filtro":cidades_filtro,
              "auto_conversa":auto_conversa,"ac_modelo":ac_modelo,
              "ac_ollama_url":ac_ollama_url,"ac_intervalo":ac_intervalo,"ac_pausa":ac_pausa}
    threading.Thread(target=_send_worker, args=(params,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/send/stop", methods=["POST"])
def send_stop():
    _send_state["should_stop"] = True
    _send_notify({"type":"running","data":False})
    return jsonify({"ok": True})


@app.route("/api/send/stream")
def send_stream():
    q = queue.Queue()
    _send_subscribers.append(q)
    def generate():
        yield f"data: {json.dumps({'type':'init','stats':dict(_send_state['stats']),'running':_send_state['running'],'progress':dict(_send_state['progress'])})}\n\n"
        for log in _send_state["logs"][-50:]:
            yield f"data: {json.dumps(log)}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield f"data: {json.dumps(msg)}\n\n"
                except queue.Empty:
                    yield "data: {\"type\":\"ping\"}\n\n"
        finally:
            try: _send_subscribers.remove(q)
            except ValueError: pass
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no","Connection":"keep-alive"})


@app.route("/api/send/status")
def send_status():
    return jsonify({"running":_send_state["running"],"stats":dict(_send_state["stats"]),
                    "progress":dict(_send_state["progress"])})


@app.route("/api/send/relatorio")
def send_relatorio_json():
    """Endpoint JSON para o relatorio mobile."""
    ultimos_logs = _send_state["logs"][-30:]
    return jsonify({
        "running":       _send_state["running"],
        "stats":         dict(_send_state["stats"]),
        "progress":      dict(_send_state["progress"]),
        "rate":          dict(_send_state.get("rate", {})),
        "logs":          ultimos_logs,
        "ts":            datetime.now().strftime("%H:%M:%S"),
    })


@app.route("/api/autoconversa/start", methods=["POST"])
def autoconversa_start():
    """Inicia o Auto Conversa em background thread."""
    global _autoconv_agente
    if _autoconv_state["running"]:
        return jsonify({"error": "Auto Conversa ja esta rodando"}), 400
    if _send_state["running"]:
        return jsonify({"error": "Envio em andamento — aguarde terminar antes de iniciar o Auto Conversa"}), 400

    data = request.json or {}
    modelo    = data.get("modelo",    "llama3")
    ol_url    = data.get("ollama_url", "http://localhost:11434")
    intervalo = int(data.get("intervalo", 30))
    pausa     = int(data.get("pausa",    5))

    _autoconv_state["running"]     = True
    _autoconv_state["should_stop"] = False
    _autoconv_state["logs"]        = []
    _autoconv_state["modelo"]      = modelo
    _autoconv_state["intervalo"]   = intervalo
    _autoconv_state["pausa"]       = pausa
    for k in _autoconv_state["stats"]:
        _autoconv_state["stats"][k] = 0

    params = {"modelo": modelo, "ollama_url": ol_url, "intervalo": intervalo, "pausa": pausa}
    threading.Thread(target=_autoconv_worker, args=(params,), daemon=True).start()
    _autoconv_notify({"type": "running", "data": True})
    return jsonify({"ok": True, "modelo": modelo, "intervalo": intervalo})


@app.route("/api/autoconversa/stop", methods=["POST"])
def autoconversa_stop():
    """Para o Auto Conversa."""
    global _autoconv_agente
    _autoconv_state["should_stop"] = True
    if _autoconv_agente:
        _autoconv_agente._rodando = False
    _autoconv_notify({"type": "running", "data": False})
    return jsonify({"ok": True})


@app.route("/api/autoconversa/status")
def autoconversa_status():
    return jsonify({
        "running":   _autoconv_state["running"],
        "modelo":    _autoconv_state.get("modelo", "llama3"),
        "intervalo": _autoconv_state.get("intervalo", 30),
        "stats":     dict(_autoconv_state["stats"]),
    })


@app.route("/api/autoconversa/stream")
def autoconversa_stream():
    """SSE — logs em tempo real do Auto Conversa."""
    q = queue.Queue()
    _autoconv_subscribers.append(q)

    def generate():
        yield f"data: {json.dumps({'type':'init','running':_autoconv_state['running'],'stats':dict(_autoconv_state['stats'])})}\n\n"
        for log in _autoconv_state["logs"][-50:]:
            yield f"data: {json.dumps(log)}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield f"data: {json.dumps(msg)}\n\n"
                except queue.Empty:
                    yield "data: {\"type\":\"ping\"}\n\n"
        finally:
            try: _autoconv_subscribers.remove(q)
            except ValueError: pass

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no",
                             "Connection": "keep-alive"})


@app.route("/api/autoconversa/config", methods=["GET", "POST"])
def autoconversa_config():
    """GET: retorna config do agente. POST: salva config."""
    cfg_path = LOGS_DIR / "agente_config.json"
    _defaults = {
        "nome_agente":          "Consultor Digital",
        "empresa":              "Agencia de Sites",
        "objetivo":             "Agendar uma conversa rapida de 15-20 minutos",
        "servicos":             (
            "Sites profissionais personalizados e rapidos\n"
            "Aparecer no Google (SEO local)\n"
            "Funil de captacao de clientes online\n"
            "Landing pages para campanhas\n"
            "Gestao de presenca digital"
        ),
        "regras_extras":        "",
        "system_prompt_custom": "",
    }
    if request.method == "GET":
        cfg = dict(_defaults)
        if cfg_path.exists():
            try:
                with open(cfg_path, encoding="utf-8") as f:
                    cfg.update(json.load(f))
            except Exception:
                pass
        return jsonify(cfg)

    # POST — salva
    data = request.json or {}
    cfg  = dict(_defaults)
    if cfg_path.exists():
        try:
            with open(cfg_path, encoding="utf-8") as f:
                cfg.update(json.load(f))
        except Exception:
            pass
    for k in _defaults:
        if k in data:
            cfg[k] = str(data[k])
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    if _autoconv_agente:
        try:
            _autoconv_agente.agente_cfg = cfg
        except Exception:
            pass
    return jsonify({"ok": True})


@app.route("/api/autoconversa/historico")
def autoconversa_historico():
    """Retorna o historico de conversas do Auto Conversa."""
    p = LOGS_DIR / "conversa_historico.json"
    if not p.exists():
        return jsonify({"conversas": [], "total": 0})

    try:
        with open(p, encoding="utf-8") as f:
            hist = json.load(f)
    except Exception:
        return jsonify({"conversas": [], "total": 0})

    log_env = _carregar_log_envios()
    autorizados = {}
    for num, info in log_env.get("whatsapp", {}).items():
        import re as _re
        nums = _re.sub(r"\D", "", num)
        if not nums.startswith("55"): nums = "55" + nums
        autorizados[nums] = info if isinstance(info, dict) else {}

    conversas = []
    for numero, h in hist.items():
        info = autorizados.get(numero, {})
        msgs = h.get("mensagens", [])
        conversas.append({
            "numero":       numero,
            "nome":         info.get("nome") or numero,
            "nicho":        info.get("nicho", ""),
            "cidade":       info.get("cidade", ""),
            "status":       h.get("status", "ativo"),
            "total_nos":    sum(1 for m in msgs if m["de"] == "nos"),
            "total_eles":   sum(1 for m in msgs if m["de"] == "eles"),
            "ultima_ts":    h.get("ultima_resposta_ts", ""),
            "ultima_msg":   msgs[-1]["texto"][:120] if msgs else "",
            "ultima_de":    msgs[-1]["de"] if msgs else "",
        })

    conversas.sort(key=lambda x: x["ultima_ts"], reverse=True)
    return jsonify({"conversas": conversas, "total": len(conversas)})


@app.route("/api/leads/historico")
def leads_historico():
    """Retorna historico completo de envios."""
    log = _carregar_log_envios()
    hoje_str = date.today().strftime("%Y-%m-%d")

    def _entrada(chave, v, canal):
        if isinstance(v, dict):
            return {"numero": chave, "canal": canal,
                    "ts": v.get("ts",""), "nome": v.get("nome",""),
                    "nicho": v.get("nicho",""), "cidade": v.get("cidade",""),
                    "hoje": str(v.get("ts","")).startswith(hoje_str)}
        return {"numero": chave, "canal": canal, "ts": str(v), "nome": "", "nicho": "", "cidade": "",
                "hoje": str(v).startswith(hoje_str)}

    historico = []
    for num, v in log.get("whatsapp", {}).items():
        historico.append(_entrada(num, v, "whatsapp"))
    for em, v in log.get("email", {}).items():
        historico.append(_entrada(em, v, "email"))

    historico.sort(key=lambda x: x["ts"], reverse=True)
    hoje_count = sum(1 for h in historico if h["hoje"] and h["canal"] == "whatsapp")
    return jsonify({"historico": historico, "total": len(historico), "hoje": hoje_count})


@app.route("/api/leads/central")
def leads_central():
    """Agrega leads de TODOS os arquivos xlsx com filtros opcionais."""
    import openpyxl as _opx

    filtro_nicho    = request.args.get("nicho",    "").lower()
    filtro_cidade   = request.args.get("cidade",   "").lower()
    filtro_prio     = request.args.get("prio",     "")
    filtro_status   = request.args.get("status",   "")
    filtro_tipo_tel = request.args.get("tipo_tel", "")
    page_num        = int(request.args.get("page", 1))
    per_page        = int(request.args.get("per", 50))

    log = _carregar_log_envios()
    ja_enviados_wa = set(log.get("whatsapp", {}).keys())

    todos = []
    arquivos = list(sorted(LEADS_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True))
    arquivos += list(sorted(BASE_DIR.glob("*.xlsx"),  key=lambda x: x.stat().st_mtime, reverse=True))
    arquivos_vistos = set()

    for path in arquivos:
        if path.name in arquivos_vistos:
            continue
        arquivos_vistos.add(path.name)
        try:
            wb = _opx.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb["Leads"]
            hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            col_nome   = hdrs.get("Nome",         1)
            col_nicho  = hdrs.get("Nicho",        2)
            col_cidade = hdrs.get("Cidade",       3)
            col_prio   = hdrs.get("Prioridade",   7)
            col_email  = hdrs.get("Email",        10)
            col_wpp    = hdrs.get("WhatsApp",     11)
            col_senv   = hdrs.get("Status Envio", None)
            col_tipo   = hdrs.get("Tipo Tel",     None)
            col_wstat  = hdrs.get("WA Status",    None)

            for row in range(2, ws.max_row + 1):
                nome = ws.cell(row, col_nome).value
                if not nome:
                    continue
                wpp_raw  = str(ws.cell(row, col_wpp  ).value or "").strip()
                nicho    = str(ws.cell(row, col_nicho ).value or "")
                cidade   = str(ws.cell(row, col_cidade).value or "")
                prio     = str(ws.cell(row, col_prio  ).value or "")
                email    = str(ws.cell(row, col_email ).value or "").strip()
                status   = str(ws.cell(row, col_senv  ).value or "") if col_senv  else ""
                tipo_tel = str(ws.cell(row, col_tipo  ).value or "") if col_tipo  else _classificar_numero(wpp_raw) if wpp_raw else ""
                wa_stat  = str(ws.cell(row, col_wstat ).value or "") if col_wstat else ""

                num_d = re.sub(r"\D", "", wpp_raw)
                if num_d and not num_d.startswith("55"):
                    num_d = "55" + num_d
                ja_enviou_wa = num_d in ja_enviados_wa if num_d else False

                if filtro_nicho  and filtro_nicho  not in nicho.lower():   continue
                if filtro_cidade and filtro_cidade not in cidade.lower():   continue
                if filtro_prio   and filtro_prio   not in prio:             continue
                if filtro_tipo_tel and filtro_tipo_tel != tipo_tel.lower(): continue
                if filtro_status == "enviado"     and not (ja_enviou_wa or "enviado" in status.lower()): continue
                if filtro_status == "nao_enviado" and (ja_enviou_wa or "enviado" in status.lower()):     continue

                todos.append({
                    "nome":      str(nome),
                    "nicho":     nicho,
                    "cidade":    cidade,
                    "prio":      prio,
                    "email":     email,
                    "wpp":       wpp_raw,
                    "tipo_tel":  tipo_tel,
                    "wa_stat":   wa_stat,
                    "status":    status,
                    "enviado":   ja_enviou_wa or "enviado" in status.lower(),
                    "arquivo":   path.name,
                })
            wb.close()
        except Exception:
            pass

    total = len(todos)
    inicio = (page_num - 1) * per_page
    pagina = todos[inicio: inicio + per_page]
    return jsonify({"leads": pagina, "total": total, "page": page_num, "per": per_page,
                    "pages": max(1, (total + per_page - 1) // per_page)})


@app.route("/api/leads/filtros")
def leads_filtros():
    """Retorna nichos, cidades e prioridades disponíveis."""
    import openpyxl as _opx
    arquivos_param = request.args.getlist("arquivo")

    nichos   = set()
    cidades  = set()
    prios    = set()
    todos_arqs = list(sorted(LEADS_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True))
    todos_arqs+= list(sorted(BASE_DIR.glob("*.xlsx"),  key=lambda x: x.stat().st_mtime, reverse=True))

    for path in todos_arqs:
        if arquivos_param and path.name not in arquivos_param:
            continue
        try:
            wb = _opx.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb["Leads"]
            hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            col_nicho  = hdrs.get("Nicho",      2)
            col_cidade = hdrs.get("Cidade",     3)
            col_prio   = hdrs.get("Prioridade", 7)
            for row in range(2, ws.max_row + 1):
                n = ws.cell(row, col_nicho ).value
                c = ws.cell(row, col_cidade).value
                p = ws.cell(row, col_prio  ).value
                if n: nichos.add(str(n))
                if c: cidades.add(str(c))
                if p: prios.add(str(p))
            wb.close()
        except Exception:
            pass

    return jsonify({
        "nichos":  sorted(nichos),
        "cidades": sorted(cidades),
        "prios":   sorted(prios),
    })


@app.route("/central")
def central_leads():
    """Pagina Central de Leads — visao unificada de todos os arquivos com filtros."""
    html = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Central de Leads</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#0f0f0f;color:#e5e5e5;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:16px}
h1{font-size:1.25rem;font-weight:700;margin-bottom:14px;color:#fff}
.filtros{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px;align-items:center}
.filtro-grp{display:flex;flex-direction:column;gap:3px}
.filtro-lbl{font-size:.7rem;color:#6b7280;text-transform:uppercase;letter-spacing:.05em}
select,input[type=text]{background:#1a1a1a;border:1px solid #333;color:#e5e5e5;
  border-radius:8px;padding:6px 10px;font-size:.82rem;min-width:140px;outline:none}
select:focus,input:focus{border-color:#25D366}
.btn{background:#25D366;color:#000;border:none;border-radius:8px;padding:7px 16px;
  font-size:.82rem;font-weight:600;cursor:pointer;display:flex;align-items:center;gap:5px}
.btn:hover{background:#1fbc58}
.btn-sec{background:#1a1a1a;color:#e5e5e5;border:1px solid #333}
.btn-sec:hover{border-color:#555}
.stats-bar{display:flex;gap:12px;margin-bottom:12px;flex-wrap:wrap}
.stat-chip{background:#1a1a1a;border:1px solid #2a2a2a;border-radius:8px;
  padding:6px 12px;font-size:.78rem;color:#9ca3af}
.stat-chip strong{color:#fff}
table{width:100%;border-collapse:collapse;font-size:.78rem}
thead th{background:#1a1a1a;color:#9ca3af;font-weight:600;padding:8px 10px;
  text-align:left;border-bottom:1px solid #2a2a2a;white-space:nowrap;font-size:.72rem;text-transform:uppercase}
tbody tr{border-bottom:1px solid #1a1a1a}
tbody tr:hover{background:#151515}
td{padding:7px 10px;vertical-align:middle}
.badge{display:inline-block;padding:2px 7px;border-radius:999px;font-size:.68rem;font-weight:600}
.b-alta{background:#1a3a1a;color:#4ade80}
.b-media{background:#1a2a3a;color:#60a5fa}
.b-baixa{background:#2a2a1a;color:#d4a843}
.b-enviado{background:#1a3a1a;color:#4ade80}
.b-nao{background:#2a1a1a;color:#f87171}
.b-fixo{background:#2a1a2a;color:#c084fc}
.b-celular{background:#1a2a2a;color:#67e8f9}
.nome-cell{font-weight:600;color:#fff;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.arquivo-cell{color:#6b7280;font-size:.68rem;max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.paginacao{display:flex;gap:8px;margin-top:14px;align-items:center;justify-content:center}
.pg-btn{background:#1a1a1a;border:1px solid #333;color:#e5e5e5;border-radius:8px;
  padding:5px 12px;font-size:.78rem;cursor:pointer}
.pg-btn:hover{border-color:#555}
.pg-btn:disabled{opacity:.4;cursor:default}
.pg-info{font-size:.78rem;color:#6b7280}
.loading{text-align:center;padding:30px;color:#6b7280}
</style>
</head>
<body>
<h1>Central de Leads</h1>
<div class="filtros">
  <div class="filtro-grp">
    <span class="filtro-lbl">Cidade</span>
    <select id="f-cidade"><option value="">Todas</option></select>
  </div>
  <div class="filtro-grp">
    <span class="filtro-lbl">Prioridade</span>
    <select id="f-prio">
      <option value="">Todas</option>
      <option value="Alta">Alta</option>
      <option value="Media">Media</option>
      <option value="Baixa">Baixa</option>
    </select>
  </div>
  <div class="filtro-grp">
    <span class="filtro-lbl">Status</span>
    <select id="f-status">
      <option value="">Todos</option>
      <option value="nao_enviado">Nao enviados</option>
      <option value="enviado">Ja enviados</option>
    </select>
  </div>
  <button class="btn" onclick="buscar(1)">Filtrar</button>
  <button class="btn btn-sec" onclick="limpar()">Limpar</button>
</div>
<div class="stats-bar" id="stats-bar"><div class="stat-chip">Carregando...</div></div>
<div id="tabela-wrap"><div class="loading">Carregando leads...</div></div>
<div class="paginacao" id="paginacao" style="display:none">
  <button class="pg-btn" id="pg-prev" onclick="trocarPg(-1)">Anterior</button>
  <span class="pg-info" id="pg-info"></span>
  <button class="pg-btn" id="pg-next" onclick="trocarPg(1)">Proxima</button>
</div>
<script>
let currentPage=1,totalPages=1;
async function carregarFiltros(){
  try{
    const r=await fetch('/api/leads/filtros');
    const d=await r.json();
    const sc=document.getElementById('f-cidade');
    sc.innerHTML='<option value="">Todas</option>';
    d.cidades.forEach(c=>{const o=document.createElement('option');o.value=c;o.textContent=c;sc.appendChild(o);});
  }catch(e){}
}
async function buscar(pg){
  currentPage=pg;
  const cidade=document.getElementById('f-cidade').value;
  const prio=document.getElementById('f-prio').value;
  const status=document.getElementById('f-status').value;
  const params=new URLSearchParams({page:pg,per:50});
  if(cidade)params.set('cidade',cidade);
  if(prio)params.set('prio',prio);
  if(status)params.set('status',status);
  document.getElementById('tabela-wrap').innerHTML='<div class="loading">Carregando...</div>';
  try{
    const r=await fetch('/api/leads/central?'+params.toString());
    const d=await r.json();
    totalPages=d.pages;
    document.getElementById('paginacao').style.display='flex';
    document.getElementById('pg-info').textContent='Pagina '+d.page+' de '+d.pages+' ('+d.total+' leads)';
    document.getElementById('pg-prev').disabled=d.page<=1;
    document.getElementById('pg-next').disabled=d.page>=d.pages;
    const env=d.leads.filter(l=>l.enviado).length;
    document.getElementById('stats-bar').innerHTML=
      '<div class="stat-chip"><strong>'+d.total+'</strong> total</div>'+
      '<div class="stat-chip"><strong>'+env+'</strong> enviados</div>';
    if(!d.leads.length){document.getElementById('tabela-wrap').innerHTML='<div class="loading">Nenhum lead encontrado.</div>';return;}
    let html='<table><thead><tr><th>Nome</th><th>Nicho</th><th>Cidade</th><th>Prio</th><th>Status</th><th>Arquivo</th></tr></thead><tbody>';
    d.leads.forEach(l=>{
      const bPrio=l.prio.includes('Alta')?'b-alta':l.prio.includes('edia')?'b-media':'b-baixa';
      const bEnv=l.enviado?'b-enviado':'b-nao';
      html+='<tr><td><span class="nome-cell" title="'+l.nome+'">'+l.nome+'</span></td>'+
        '<td>'+l.nicho+'</td><td>'+l.cidade+'</td>'+
        '<td><span class="badge '+bPrio+'">'+l.prio+'</span></td>'+
        '<td><span class="badge '+bEnv+'">'+(l.enviado?'enviado':'pendente')+'</span></td>'+
        '<td><span class="arquivo-cell">'+l.arquivo+'</span></td></tr>';
    });
    html+='</tbody></table>';
    document.getElementById('tabela-wrap').innerHTML=html;
  }catch(e){document.getElementById('tabela-wrap').innerHTML='<div class="loading">Erro: '+e.message+'</div>';}
}
function trocarPg(d){buscar(Math.max(1,Math.min(totalPages,currentPage+d)));}
function limpar(){document.getElementById('f-cidade').value='';document.getElementById('f-prio').value='';document.getElementById('f-status').value='';buscar(1);}
carregarFiltros();buscar(1);
</script>
</body>
</html>"""
    return html


@app.route("/relatorio")
def relatorio_mobile():
    """Pagina mobile-friendly de relatorio."""
    html = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Relatorio de Envios</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:#0f0f0f;color:#e5e5e5;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:16px;min-height:100vh}
  h1{font-size:1.3rem;font-weight:700;text-align:center;margin-bottom:16px;color:#fff}
  .cards{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px}
  .card{background:#1a1a1a;border:1px solid #2a2a2a;border-radius:12px;padding:14px;text-align:center}
  .card-val{font-size:2rem;font-weight:700;line-height:1}
  .card-lbl{font-size:.72rem;color:#9ca3af;margin-top:4px;text-transform:uppercase}
  .card-wa .card-val{color:#25D366}
  .card-email .card-val{color:#60a5fa}
  .card-err .card-val{color:#f87171}
  .card-total .card-val{color:#f59e0b}
  .logs-wrap{background:#1a1a1a;border:1px solid #2a2a2a;border-radius:12px;padding:12px;max-height:340px;overflow-y:auto;margin-bottom:12px}
  .logs-title{font-size:.78rem;color:#6b7280;text-transform:uppercase;margin-bottom:8px}
  .log-line{font-size:.78rem;line-height:1.5;padding:2px 0;border-bottom:1px solid #1f1f1f;display:flex;gap:6px}
  .log-ts{color:#4b5563;flex-shrink:0;font-size:.7rem;padding-top:1px}
  .log-msg{word-break:break-word}
  .lvl-success{color:#4ade80}.lvl-error{color:#f87171}.lvl-warn{color:#fbbf24}.lvl-info{color:#e5e5e5}
  .refresh-ts{text-align:center;font-size:.72rem;color:#4b5563;margin-top:12px}
</style>
</head>
<body>
<h1>Relatorio de Envios</h1>
<div class="cards">
  <div class="card card-total"><div class="card-val" id="c-total">-</div><div class="card-lbl">Total</div></div>
  <div class="card card-wa"><div class="card-val" id="c-wa">-</div><div class="card-lbl">WhatsApp</div></div>
  <div class="card card-email"><div class="card-val" id="c-email">-</div><div class="card-lbl">Email</div></div>
  <div class="card card-err"><div class="card-val" id="c-err">-</div><div class="card-lbl">Erros</div></div>
</div>
<div class="logs-wrap">
  <div class="logs-title">Ultimos logs</div>
  <div id="logs-container"></div>
</div>
<div class="refresh-ts">Atualiza a cada 5s — <span id="last-ts">-</span></div>
<script>
const lvlClass={success:'lvl-success',error:'lvl-error',warn:'lvl-warn',info:'lvl-info'};
async function refresh(){
  try{
    const r=await fetch('/api/send/relatorio');
    const d=await r.json();
    const st=d.stats||{};
    document.getElementById('c-total').textContent=(st.whatsapp||0)+(st.emails||0);
    document.getElementById('c-wa').textContent=st.whatsapp||0;
    document.getElementById('c-email').textContent=st.emails||0;
    document.getElementById('c-err').textContent=st.errors||0;
    const logs=d.logs||[];
    const cont=document.getElementById('logs-container');
    cont.innerHTML='';
    for(let i=logs.length-1;i>=0;i--){
      const l=logs[i];
      const div=document.createElement('div');
      div.className='log-line';
      div.innerHTML='<span class="log-ts">'+l.ts+'</span><span class="log-msg '+(lvlClass[l.level]||'lvl-info')+'">'+l.msg+'</span>';
      cont.appendChild(div);
    }
    document.getElementById('last-ts').textContent=d.ts;
  }catch(e){document.getElementById('last-ts').textContent='erro: '+e.message;}
}
refresh();
setInterval(refresh,5000);
</script>
</body>
</html>"""
    return html


# Classificacao inline de numero brasileiro
def _classificar_numero(numero: str) -> str:
    d = re.sub(r"\D", "", numero)
    if d.startswith("55"):
        d = d[2:]
    if len(d) == 11 and d[2] == "9":
        return "celular"
    if len(d) == 10 and d[2] in "2345":
        return "fixo"
    if len(d) == 11 and d[2] in "2345":
        return "fixo"
    if len(d) in (10, 11):
        return "celular"
    return "desconhecido"


# Log persistente de envios (anti-duplicata cross-session)
_ENVIOS_LOG_FILE = LOGS_DIR / "enviados_log.json"


def _carregar_log_envios() -> dict:
    """Carrega o registro de envios anteriores."""
    try:
        if _ENVIOS_LOG_FILE.exists():
            with open(_ENVIOS_LOG_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {"whatsapp": {}, "email": {}}


def _salvar_log_envios(log: dict):
    """Persiste o log de envios no disco."""
    try:
        with open(_ENVIOS_LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# Playwright WA Sender
class _PlaywrightWASender:
    _PROFILE = BASE_DIR / "wa_profile"

    def __init__(self, log_fn=None):
        self._log   = log_fn or (lambda m, *_: None)
        self._pw    = None
        self._ctx   = None
        self._page  = None
        self.pronto = False

    def iniciar(self) -> bool:
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            return False

        self._PROFILE.mkdir(exist_ok=True)
        self._pw  = sync_playwright().__enter__()

        self._ctx = self._pw.chromium.launch_persistent_context(
            user_data_dir=str(self._PROFILE),
            headless=False,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
            viewport={"width": 1280, "height": 800},
        )
        self._page = self._ctx.pages[0] if self._ctx.pages else self._ctx.new_page()
        self._page.set_default_timeout(30000)

        if_first = not (self._PROFILE / "Default" / "Cookies").exists()
        self._log("Abrindo WhatsApp Web para envio...", "info")
        if if_first:
            self._log("Primeiro uso — escaneie o QR code (ate 90s)...", "warn")
        else:
            self._log("Perfil salvo detectado — tentando login automatico...", "info")

        self._page.goto("https://web.whatsapp.com", wait_until="domcontentloaded", timeout=30000)

        try:
            self._page.wait_for_selector(
                '#pane-side, div[aria-label="Lista de conversas"], div[aria-label="Chat list"]',
                timeout=90000,
            )
            self._log("WhatsApp Web pronto para envio!", "success")
            self.pronto = True
            return True
        except Exception:
            self._log("Timeout login WA — WhatsApp desativado.", "error")
            return False

    def enviar(self, numero: str, mensagem: str) -> dict:
        import urllib.parse as _up
        page = self._page

        nums = re.sub(r"\D", "", numero)
        if not nums.startswith("55"):
            nums = "55" + nums

        url = (
            f"https://web.whatsapp.com/send?phone={nums}"
            f"&text={_up.quote(mensagem, safe='')}"
            f"&type=phone_number&app_absent=0"
        )

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            time.sleep(2)
        except Exception as e:
            return {"ok": False, "motivo": f"nav: {str(e)[:50]}"}

        for sel in [
            'div[data-testid="popup-controls"]',
            'div[data-animate-modal-body]',
            'div[role="dialog"]',
        ]:
            try:
                popup = page.locator(sel).first
                if popup.is_visible(timeout=4000):
                    txt = popup.inner_text(timeout=2000).lower()
                    invalido = any(w in txt for w in [
                        "invalid", "invalido", "not found", "nao encontrado",
                        "phone number shared", "not registered",
                    ])
                    if invalido:
                        for btn in [
                            'button[data-testid="popup-controls-ok"]',
                            'button:has-text("OK")', 'button:has-text("Ok")',
                        ]:
                            try:
                                page.locator(btn).first.click(timeout=1500)
                                break
                            except Exception:
                                pass
                        return {"ok": False, "motivo": "sem_whatsapp"}
            except Exception:
                pass

        _SELS_INPUT = [
            'div[contenteditable="true"][data-tab]',
            'div[aria-label="Digite uma mensagem"]',
            'div[aria-label="Type a message"]',
            'div[data-testid="conversation-compose-box-input"]',
        ]
        chat_carregou = False
        sel_input_usado = None
        for sel in _SELS_INPUT:
            try:
                page.wait_for_selector(sel, timeout=10000)
                chat_carregou = True
                sel_input_usado = sel
                break
            except Exception:
                pass

        if not chat_carregou:
            try:
                page.wait_for_selector('footer[data-testid="conversation-footer"]', timeout=5000)
                chat_carregou = True
            except Exception:
                pass

        if not chat_carregou:
            return {"ok": False, "motivo": "timeout_chat"}

        _SELS_BTN_SEND = [
            'button[data-testid="send"]',
            'button[aria-label="Enviar mensagem"]',
            'button[aria-label="Send message"]',
            '[data-testid="send"]',
        ]
        clicou_enviar = False
        for sel in _SELS_BTN_SEND:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=2000):
                    btn.click(timeout=2000)
                    clicou_enviar = True
                    break
            except Exception:
                pass

        if not clicou_enviar:
            try:
                page.keyboard.press("Enter")
                clicou_enviar = True
            except Exception as e:
                return {"ok": False, "motivo": f"erro_envio: {str(e)[:40]}"}

        time.sleep(2)
        self._log("Mensagem enviada.", "warn")
        return {"ok": True, "motivo": "sem_confirmacao_visual"}

    def fechar(self):
        try:
            self._ctx.close()
        except Exception:
            pass
        try:
            self._pw.__exit__(None, None, None)
        except Exception:
            pass


# Auto Conversa Worker
def _autoconv_worker(params: dict):
    global _autoconv_agente

    mod = _load_mod("autoconversa", BASE_DIR / "core" / "whatsapp.py")
    if not mod:
        _autoconv_log("Nao carregou core/whatsapp.py", "error")
        _autoconv_state["running"] = False
        _autoconv_notify({"type": "running", "data": False})
        return

    modelo   = params.get("modelo",   "llama3")
    ol_url   = params.get("ollama_url", "http://localhost:11434")
    intervalo = params.get("intervalo", 30)
    pausa     = params.get("pausa",    5)

    if not mod.verificar_ollama(ol_url, modelo):
        _autoconv_log(
            f"Ollama nao disponivel em {ol_url}. "
            f"Instale em https://ollama.ai/ e execute: ollama pull {modelo}", "error")
        _autoconv_state["running"] = False
        _autoconv_notify({"type": "running", "data": False})
        return

    autorizados = mod.carregar_autorizados()
    historico   = mod.carregar_historico()

    if not autorizados:
        _autoconv_log("Nenhum numero autorizado em enviados_log.json.", "error")
        _autoconv_state["running"] = False
        _autoconv_notify({"type": "running", "data": False})
        return

    agente = mod.AutoConversador(
        modelo=modelo,
        ollama_url=ol_url,
        intervalo=intervalo,
        pausa_resposta=pausa,
    )
    agente.autorizados = autorizados
    agente.historico   = historico
    _autoconv_agente   = agente

    orig = sys.stdout

    class _ACWriter:
        def write(self_, t):
            if t and t.strip():
                s = t.strip()
                lvl = "info"
                if any(x in s for x in ("ERRO", "erro", "falhou", "Erro")): lvl = "error"
                elif any(x in s for x in ("Aviso", "aviso", "Bot")): lvl = "warn"
                elif any(x in s for x in ("enviada", "Resposta", "pronto")): lvl = "success"
                elif any(x in s for x in ("Aguard", "Ciclo", "===")): lvl = "dim"
                _autoconv_log(s, lvl)
                if "respondido" in s.lower():
                    _autoconv_state["stats"]["respondidos"] += 1
                elif "automatica" in s.lower() or "bot" in s.lower():
                    _autoconv_state["stats"]["bots_ignorados"] += 1
                elif "sem interesse" in s.lower():
                    _autoconv_state["stats"]["sem_interesse"] += 1
                _autoconv_notify({"type": "stats", "data": dict(_autoconv_state["stats"])})
        def flush(self_): pass

    sys.stdout = _ACWriter()

    try:
        _autoconv_log(
            f"Auto Conversa iniciando — {len(autorizados)} contatos autorizados | "
            f"modelo: {modelo} | intervalo: {intervalo}s", "success")
        agente._iniciar_whatsapp()
        agente.rodar()
    except Exception as e:
        _autoconv_log(f"Erro fatal: {e}", "error")
    finally:
        sys.stdout = orig
        try:
            agente.fechar()
        except Exception:
            pass
        mod.salvar_historico(agente.historico)
        _autoconv_state["running"]      = False
        _autoconv_state["should_stop"]  = False
        _autoconv_agente = None
        _autoconv_notify({"type": "running", "data": False})
        _autoconv_log("Auto Conversa encerrada.", "warn")


# Send Worker
def _send_worker(params: dict):
    import base64
    from email.mime.text import MIMEText

    files          = params["files"]
    channels       = params["channels"]
    project        = params["project"]
    priorities     = params["priorities"]
    max_sends      = params["max_sends"]
    templates      = params["templates"]
    limite_diario  = params.get("limite_diario",  50)
    delay_min      = params.get("delay_min",       20)
    delay_max      = params.get("delay_max",       45)
    pausa_a_cada   = params.get("pausa_a_cada",    10)
    pausa_duracao  = params.get("pausa_duracao",  300)
    nichos_filtro  = [n.lower() for n in params.get("nichos_filtro",  [])]
    cidades_filtro = [c.lower() for c in params.get("cidades_filtro", [])]

    envios_log = _carregar_log_envios()
    ja_enviados_wa    = set(envios_log.get("whatsapp", {}).keys())
    ja_enviados_email = set(envios_log.get("email",    {}).keys())

    hoje_str = date.today().strftime("%Y-%m-%d")

    def _ts_de_entrada(v) -> str:
        if isinstance(v, dict):
            return str(v.get("ts", ""))
        return str(v)

    enviados_hoje = sum(
        1 for v in envios_log.get("whatsapp", {}).values()
        if _ts_de_entrada(v).startswith(hoje_str)
    )
    _send_state["rate"]["enviados_hoje"] = enviados_hoje
    _send_notify({"type": "rate", "data": dict(_send_state["rate"])})

    if enviados_hoje >= limite_diario:
        _send_log(
            f"Limite diario atingido ({enviados_hoje}/{limite_diario} msgs). "
            f"Aguarde ate amanha para continuar.", "error")
        _send_state["running"] = False
        _send_notify({"type": "running", "data": False})
        return

    _send_log(
        f"Anti-ban: {enviados_hoje}/{limite_diario} msgs hoje | "
        f"delay {delay_min}-{delay_max}s | pausa {pausa_duracao//60}min a cada {pausa_a_cada} msgs",
        "info")

    envios_sessao_wa = 0

    def _registrar_envio(canal: str, chave: str, nome: str = "", nicho: str = "", cidade: str = ""):
        nonlocal enviados_hoje, envios_sessao_wa
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        entrada = {"ts": ts, "nome": nome, "nicho": nicho, "cidade": cidade}
        envios_log.setdefault(canal, {})[chave] = entrada
        if canal == "whatsapp":
            ja_enviados_wa.add(chave)
            enviados_hoje     += 1
            envios_sessao_wa  += 1
            _send_state["rate"]["enviados_hoje"] = enviados_hoje
            _send_notify({"type": "rate", "data": dict(_send_state["rate"])})
        else:
            ja_enviados_email.add(chave)
        _salvar_log_envios(envios_log)

    def _aplicar_rate_limit_pos_wa():
        nonlocal envios_sessao_wa

        if enviados_hoje >= limite_diario:
            _send_log(
                f"Limite diario atingido ({enviados_hoje}/{limite_diario}). "
                f"Parando para proteger seu numero.", "error")
            _send_state["should_stop"] = True
            return

        if pausa_a_cada > 0 and envios_sessao_wa > 0 and envios_sessao_wa % pausa_a_cada == 0:
            fim = time.time() + pausa_duracao
            fim_str = datetime.fromtimestamp(fim).strftime("%H:%M:%S")
            _send_state["rate"]["pausando_ate"] = fim_str
            _send_notify({"type": "rate", "data": dict(_send_state["rate"])})
            _send_log(
                f"Pausa de {pausa_duracao//60}min apos {envios_sessao_wa} msgs "
                f"(anti-ban) — retorna as {fim_str}", "warn")
            while time.time() < fim:
                if _send_state["should_stop"]:
                    break
                time.sleep(5)
            _send_state["rate"]["pausando_ate"] = ""
            _send_notify({"type": "rate", "data": dict(_send_state["rate"])})
            if not _send_state["should_stop"]:
                _send_log("Pausa terminada — retomando envios.", "info")
            return

        delay = random.uniform(delay_min, delay_max)
        _send_log(f"Aguardando {delay:.0f}s (anti-ban)...", "dim")
        time.sleep(delay)

    if project == "both":
        proj_url  = f"{PROJECT_URLS['xiiter']} | {PROJECT_URLS['inserth']}"
        proj_name = "Xiiter / Inserth"
    else:
        proj_url  = PROJECT_URLS.get(project, PROJECT_URLS["xiiter"])
        proj_name = PROJECT_NAMES.get(project, "Xiiter")

    do_email    = "email"    in channels
    do_whatsapp = "whatsapp" in channels

    wa_sender = None
    pywhatkit = None

    if do_whatsapp:
        try:
            import playwright  # noqa: F401
            wa_sender = _PlaywrightWASender(log_fn=_send_log)
            if not wa_sender.iniciar():
                wa_sender = None
        except ImportError:
            pass

        if wa_sender is None:
            try:
                import pywhatkit as pwk
                pywhatkit = pwk
                _send_log("pywhatkit carregado (fallback — sem verificacao real).", "warn")
            except ImportError:
                _send_log("Nenhum cliente WA disponivel.", "warn")
                do_whatsapp = False
        else:
            _send_log("Playwright WA pronto.", "success")

    gmail_service    = None
    gmail_remetente  = ""
    if do_email:
        try:
            spec = importlib.util.spec_from_file_location("enviar", BASE_DIR / "core" / "email_sender.py")
            mod  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            gmail_service, gmail_remetente = mod.autenticar_gmail()
            _send_log(f"Gmail: {gmail_remetente}", "success")
        except Exception as e:
            _send_log(f"Erro Gmail: {e}", "error")
            do_email = False

    if not do_email and not do_whatsapp:
        _send_log("Nenhum canal disponivel. Abortando.", "error")
        _send_state["running"] = False
        _send_notify({"type":"running","data":False})
        return

    total_enviados = 0

    def _fmt_msg(template: str, nome: str, nicho: str, cidade: str) -> str:
        return (template
                .replace("{nome}",    nome)
                .replace("{nicho}",   nicho)
                .replace("{cidade}",  cidade)
                .replace("{projeto}", proj_name)
                .replace("{url}",     proj_url))

    for filepath_str in files:
        if _send_state["should_stop"]: break

        path = Path(filepath_str) if os.path.isabs(filepath_str) else LEADS_DIR / filepath_str
        if not path.exists(): path = BASE_DIR / filepath_str
        if not path.exists():
            _send_log(f"Nao encontrado: {filepath_str}", "warn"); continue

        _send_log(f"{path.name}", "info")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.load_workbook(str(path))
            ws = wb["Leads"]
        except Exception as e:
            _send_log(f"Erro abrindo {path.name}: {e}", "error"); continue

        hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        col_nome      = hdrs.get("Nome",       1)
        col_nicho     = hdrs.get("Nicho",      2)
        col_cidade    = hdrs.get("Cidade",     3)
        col_email     = hdrs.get("Email",      10)
        col_wpp       = hdrs.get("WhatsApp",   11)
        col_prio      = hdrs.get("Prioridade", 7)
        col_senvio    = hdrs.get("Status Envio", None)
        col_tipo_tel  = hdrs.get("Tipo Tel",   None)
        col_wa_status = hdrs.get("WA Status",  None)

        if not col_senvio:
            col_senvio = ws.max_column + 1
            cell = ws.cell(1, col_senvio, value="Status Envio")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1F4E79")

        for row in range(2, ws.max_row + 1):
            if _send_state["should_stop"] or total_enviados >= max_sends: break
            nome = ws.cell(row, col_nome).value
            if not nome: continue
            prio = ws.cell(row, col_prio).value or ""
            if not any(p in prio for p in priorities): continue

            nicho_row  = str(ws.cell(row, col_nicho ).value or "").lower()
            cidade_row = str(ws.cell(row, col_cidade).value or "").lower()
            if nichos_filtro  and not any(n in nicho_row  for n in nichos_filtro):  continue
            if cidades_filtro and not any(c in cidade_row for c in cidades_filtro): continue

            status_env = str(ws.cell(row, col_senvio).value or "").lower()
            if "enviado" in status_env:
                continue

            email_dest = str(ws.cell(row, col_email).value or "").strip()
            wpp_raw    = str(ws.cell(row, col_wpp).value   or "").strip()
            nicho      = str(ws.cell(row, col_nicho).value  or "")
            cidade     = str(ws.cell(row, col_cidade).value or "")
            nome_str   = str(nome)

            num_digits = re.sub(r"\D", "", wpp_raw)
            if num_digits and not num_digits.startswith("55"):
                num_digits = "55" + num_digits
            num_intl = ("+" + num_digits) if num_digits else ""

            if col_tipo_tel:
                tipo_tel = str(ws.cell(row, col_tipo_tel).value or "").lower()
            else:
                tipo_tel = _classificar_numero(wpp_raw) if wpp_raw else "desconhecido"
            wa_status = str(ws.cell(row, col_wa_status).value or "").lower() if col_wa_status else ""

            _send_state["stats"]["total"] += 1
            _send_notify({"type":"stats","data":dict(_send_state["stats"])})

            sent_ch = []

            if do_email and email_dest and "@" in email_dest:
                email_key = email_dest.lower()
                if email_key in ja_enviados_email:
                    _send_log(f"Email ja enviado anteriormente: {nome_str}", "dim")
                else:
                    try:
                        subj = _fmt_msg(
                            templates.get("email_subject") or f"Proposta digital para {nome_str}",
                            nome_str, nicho, cidade)
                        body = _fmt_msg(
                            templates.get("email_body") or
                            f"Ola,\n\nVi {nome_str} no Google Maps e quero apresentar o {proj_name}.\n\nAcesse: {proj_url}\n\nPode ser exatamente o que precisam para crescer digitalmente em {cidade}.\n\nQualquer duvida, estou a disposicao.",
                            nome_str, nicho, cidade)
                        msg_mime = MIMEText(body, "plain", "utf-8")
                        msg_mime["to"] = email_dest
                        msg_mime["from"] = gmail_remetente
                        msg_mime["subject"] = subj
                        raw = base64.urlsafe_b64encode(msg_mime.as_bytes()).decode()
                        gmail_service.users().messages().send(userId="me", body={"raw": raw}).execute()
                        sent_ch.append("email")
                        _send_state["stats"]["emails"] += 1
                        _send_log(f"Email {nome_str} -> {email_dest}", "success")
                        _registrar_envio("email", email_key, nome=nome_str, nicho=nicho, cidade=cidade)
                        time.sleep(3 + random.uniform(0, 2))
                    except Exception as e:
                        _send_log(f"Erro Email ({nome_str}): {e}", "error")
                        _send_state["stats"]["errors"] += 1

            if do_whatsapp and num_intl:
                wa_key = num_digits

                if wa_key in ja_enviados_wa:
                    _send_log(f"WA ja enviado anteriormente: {nome_str} ({num_intl})", "dim")
                elif tipo_tel == "fixo":
                    _send_log(f"Pulando WA — telefone fixo: {nome_str}", "dim")
                elif wa_status in ("sem_whatsapp", "bot_ignorado", "inacessivel"):
                    _send_log(f"Pulando WA ({wa_status}): {nome_str}", "dim")
                else:
                    wpp_msg = _fmt_msg(
                        templates.get("wpp_body") or
                        f"Ola! Vi o {nome_str} no Google Maps e queria apresentar o {proj_name}. Pode ser muito util para voces: {proj_url}",
                        nome_str, nicho, cidade)

                    if wa_sender:
                        try:
                            res = wa_sender.enviar(num_intl, wpp_msg)

                            if res["ok"]:
                                sent_ch.append("whatsapp")
                                _send_state["stats"]["whatsapp"] += 1
                                aviso = " sem confirmacao visual" if res["motivo"] == "sem_confirmacao_visual" else ""
                                _send_log(f"WA {nome_str} -> {num_intl}{aviso}", "success")
                                _registrar_envio("whatsapp", wa_key, nome=nome_str, nicho=nicho, cidade=cidade)
                                _aplicar_rate_limit_pos_wa()
                                if _send_state["should_stop"]:
                                    break
                            elif res["motivo"] == "sem_whatsapp":
                                _send_log(f"{nome_str}: numero nao esta no WhatsApp — pulando", "warn")
                                _send_state["stats"]["errors"] += 1
                                if col_wa_status:
                                    ws.cell(row, col_wa_status).value = "sem_whatsapp"
                                    from openpyxl.styles import PatternFill as _PF
                                    ws.cell(row, col_wa_status).fill = _PF("solid", start_color="F8D7DA")
                                try:
                                    wb.save(str(path))
                                except Exception:
                                    pass
                            else:
                                _send_log(f"WA falhou ({nome_str}): {res['motivo']}", "error")
                                _send_state["stats"]["errors"] += 1
                        except Exception as e:
                            _send_log(f"WhatsApp Playwright ({nome_str}): {e}", "error")
                            _send_state["stats"]["errors"] += 1

                    elif pywhatkit:
                        try:
                            pywhatkit.sendwhatmsg_instantly(
                                phone_no=num_intl, message=wpp_msg,
                                wait_time=15, tab_close=True, close_time=4)
                            sent_ch.append("whatsapp")
                            _send_state["stats"]["whatsapp"] += 1
                            _send_log(f"WA {nome_str} -> {num_intl} (sem verificacao)", "warn")
                            _registrar_envio("whatsapp", wa_key, nome=nome_str, nicho=nicho, cidade=cidade)
                            _aplicar_rate_limit_pos_wa()
                            if _send_state["should_stop"]:
                                break
                        except Exception as e:
                            _send_log(f"WhatsApp pywhatkit ({nome_str}): {e}", "error")
                            _send_state["stats"]["errors"] += 1

            if sent_ch:
                total_enviados += 1
                ts_str = datetime.now().strftime("%d/%m/%Y %H:%M")
                ws.cell(row, col_senvio).value = f"enviado ({', '.join(sent_ch)}) {ts_str}"
                ws.cell(row, col_senvio).fill  = PatternFill("solid", start_color="D4EDDA")

                try:
                    wb.save(str(path))
                except Exception as e:
                    _send_log(f"Erro ao salvar {path.name}: {e}", "warn")

                pct = int(total_enviados / max(max_sends, 1) * 100)
                _send_state["progress"] = {
                    "current": total_enviados, "total": max_sends,
                    "percent": pct, "label": nome_str}
                _send_notify({"type":"progress","data":dict(_send_state["progress"])})
                _send_notify({"type":"stats","data":dict(_send_state["stats"])})

        try: wb.close()
        except Exception: pass

    if wa_sender:
        try:
            wa_sender.fechar()
        except Exception:
            pass

    _send_state["running"]     = False
    _send_state["should_stop"] = False
    _send_notify({"type":"running","data":False})
    _send_log(
        f"Concluido — {total_enviados} enviados | "
        f"{_send_state['stats']['emails']} emails | "
        f"{_send_state['stats']['whatsapp']} WhatsApp | "
        f"{_send_state['stats']['errors']} erros.",
        "success")

    if params.get("auto_conversa") and not _autoconv_state["running"]:
        _send_log("Iniciando Auto Conversa automaticamente...", "info")
        _autoconv_state["running"]     = True
        _autoconv_state["should_stop"] = False
        _autoconv_state["logs"]        = []
        for k in _autoconv_state["stats"]: _autoconv_state["stats"][k] = 0
        ac_params = {
            "modelo":     params.get("ac_modelo",   "llama3"),
            "ollama_url": params.get("ac_ollama_url", "http://localhost:11434"),
            "intervalo":  params.get("ac_intervalo", 30),
            "pausa":      params.get("ac_pausa",     5),
        }
        threading.Thread(target=_autoconv_worker, args=(ac_params,), daemon=True).start()
        _autoconv_notify({"type": "running", "data": True})


# Worker
def _worker(params: dict):
    import importlib.util as ilu
    import subprocess

    spec    = ilu.spec_from_file_location("scraper", BASE_DIR / "core" / "scraper.py")
    scraper = ilu.module_from_spec(spec)
    scraper._GUI_MODE = True
    spec.loader.exec_module(scraper)

    orig = sys.stdout

    class _Writer:
        def write(self_, t):
            if t.strip():
                s = t.strip()
                if any(x in s for x in ("leads salvos","salvo")):
                    lvl = "success"
                elif "ERRO" in s or "erro" in s or "Erro" in s:
                    lvl = "error"
                elif any(x in s for x in ("Aviso","Duplicado","pulando","interrompido")):
                    lvl = "warn"
                elif any(x in s for x in ("---","===","Buscando","Max:")):
                    lvl = "dim"
                else:
                    lvl = "info"
                _log(s, lvl)
        def flush(self_): pass

    sys.stdout = _Writer()

    cidades           = params["cidades"];   nichos      = params["nichos"]
    max_leads         = params["max_leads"]; auto_save   = params["auto_save"]
    buscar_cnpj       = params["cnpj"];      precos      = params["precos"]
    gerar_email       = params["gerar_email"]; moeda     = params["moeda"]
    retype             = params.get("retype", False)
    retype_prioridades = params.get("retype_prioridades", ["Alta"])
    total  = len(cidades) * len(nichos)
    feitos = 0

    def _on_lead(lead):
        if retype and retype_prioridades:
            if not any(p in lead.get("prioridade", "") for p in retype_prioridades):
                try:
                    scraper._g["leads"].remove(lead)
                except (ValueError, AttributeError):
                    pass
                return

        retype_count[0] += 1
        if retype_count[0] >= max_leads:
            scraper._g["interrompido"] = True

        p = int(precos.get(nicho_atual[0], PRECOS_DEFAULT.get(nicho_atual[0], 2000)))
        lead["preco"] = f"R$ {p:,.0f}".replace(",",".") if moeda == "BRL" else f"${p}"
        st = lead.get("site_tipo","")
        _state["stats"]["total"]   += 1
        if st == "sem_site":       _state["stats"]["sem_site"] += 1
        elif st == "apenas_social":_state["stats"]["social"]   += 1
        else:                      _state["stats"]["site"]     += 1
        if lead.get("whatsapp"):   _state["stats"]["wpp"]      += 1
        if lead.get("cnpj"):       _state["stats"]["cnpj"]     += 1
        if lead.get("email"):      _state["stats"]["email"]    += 1
        if lead.get("instagram"):  _state["stats"]["insta"]    += 1
        _push_stats()

    nicho_atual  = [""]
    retype_count = [0]

    try:
        for nicho in nichos:
            nicho_atual[0] = nicho
            for cidade in cidades:
                if _state["should_stop"]:
                    _log("Scraping interrompido pelo usuario.", "warn")
                    break

                _log(f"[{feitos+1}/{total}]  {nicho}  ->  {cidade}", "info")
                _state["progress"]["label"] = f"{nicho} em {cidade}"
                _push_progress()

                safe = lambda s: re.sub(r"[^\w]", "_", s)
                ts   = datetime.now().strftime("%Y%m%d_%H%M")
                out  = str(LEADS_DIR / f"{safe(nicho)}_{safe(cidade)}_{ts}.xlsx")

                EXPANSOES_RETYPE = [
                    "",
                    " Zona Norte",
                    " Zona Sul",
                    " Zona Leste",
                    " Zona Oeste",
                    " Centro",
                    " regiao metropolitana",
                ]
                scraper_max = min(max_leads * 4, 200) if (retype and retype_prioridades) else max_leads

                if retype and retype_prioridades:
                    collected_leads = []
                    seen_keys       = set()

                    for exp_idx, expansao in enumerate(EXPANSOES_RETYPE):
                        if _state["should_stop"]:
                            break
                        if len(collected_leads) >= max_leads:
                            break

                        cidade_busca = cidade + expansao

                        if exp_idx > 0:
                            _log(
                                f"Expandindo busca -> '{cidade_busca}' "
                                f"({len(collected_leads)}/{max_leads} coletados)",
                                "info",
                            )
                            _state["progress"]["label"] = f"{nicho} em {cidade_busca}"
                            _push_progress()

                        retype_count[0]            = len(collected_leads)
                        scraper._g["interrompido"] = False
                        scraper._g["leads"]        = []
                        scraper._g["arquivo"]      = ""

                        try:
                            batch = scraper.scrape_google_maps(
                                nicho=nicho, cidade=cidade_busca,
                                max_leads=scraper_max,
                                buscar_cnpj_flag=buscar_cnpj,
                                salvar_a_cada=auto_save,
                                on_lead=_on_lead,
                            )
                        except Exception as e:
                            _log(f"Erro na expansao '{cidade_busca}': {e}", "error")
                            continue

                        if _state["should_stop"]:
                            collected_leads += [
                                l for l in batch
                                if (l["nome"].lower().strip(),
                                    l.get("endereco","")[:25].lower())
                                not in seen_keys
                            ]
                            break

                        novos = 0
                        for lead in batch:
                            key = (lead["nome"].lower().strip(),
                                   lead.get("endereco", "")[:25].lower())
                            if key not in seen_keys:
                                seen_keys.add(key)
                                collected_leads.append(lead)
                                novos += 1
                                if len(collected_leads) >= max_leads:
                                    break

                        _log(
                            f"+{novos} novos nesta busca | "
                            f"acumulado: {len(collected_leads)}/{max_leads}",
                            "info",
                        )

                    leads = collected_leads[:max_leads]

                else:
                    scraper._g["interrompido"] = False
                    scraper._g["leads"]        = []
                    scraper._g["arquivo"]      = out
                    retype_count[0]            = 0

                    try:
                        leads = scraper.scrape_google_maps(
                            nicho=nicho, cidade=cidade, max_leads=max_leads,
                            buscar_cnpj_flag=buscar_cnpj, salvar_a_cada=auto_save,
                            on_lead=_on_lead,
                        )
                    except Exception as e:
                        _log(f"Erro: {e}", "error")
                        feitos += 1
                        continue

                if _state["should_stop"]:
                    if leads:
                        scraper.salvar_xlsx(leads, out)
                    break

                if leads:
                    if retype and retype_prioridades:
                        label = " + ".join(retype_prioridades)
                        _log(f"Retype ({label}): {len(leads)} leads coletados no total", "info")

                    scraper.salvar_xlsx(leads, out)
                    _log(f"{len(leads)} leads salvos: {Path(out).name}", "success")
                    _notify({"type":"file_added"})

                    if gerar_email:
                        _log(f"Gerando emails para {Path(out).name}...", "info")
                        try:
                            r = subprocess.run(
                                [sys.executable, str(BASE_DIR/"core"/"email_generator.py"),
                                 "--leads", out, "--fallback"],
                                capture_output=True, text=True, encoding="utf-8", timeout=300
                            )
                            if r.stdout:
                                for line in r.stdout.split("\n"):
                                    if line.strip():
                                        _log(line.strip(), "info")
                        except Exception as e:
                            _log(f"Erro gerador email: {e}", "error")
                else:
                    try:
                        Path(out).unlink(missing_ok=True)
                    except Exception:
                        pass
                    _log(f"Nenhum lead para {nicho} em {cidade}.", "warn")

                feitos += 1
                _state["progress"] = {"current":feitos,"total":total,
                                      "percent":int(feitos/total*100),"label":""}
                _push_progress()

            else:
                continue
            break

    finally:
        sys.stdout = orig
        _state["running"]     = False
        _state["should_stop"] = False
        _push_stats()
        _push_progress()
        _notify({"type":"running","data":False})
        _log(f"Concluido. {feitos}/{total} buscas  |  {_state['stats']['total']} leads totais.", "success")


# CRM
def _carregar_crm() -> dict:
    try:
        if _CRM_FILE and _CRM_FILE.exists():
            with open(_CRM_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _salvar_crm(crm: dict):
    try:
        with open(_CRM_FILE, "w", encoding="utf-8") as f:
            json.dump(crm, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

@app.route("/api/crm", methods=["GET"])
def crm_get():
    """Retorna todos os leads com status CRM."""
    import openpyxl as _opx
    crm_data = _carregar_crm()
    log = _carregar_log_envios()

    leads = []
    arquivos = list(sorted(LEADS_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True))

    for path in arquivos[:20]:
        try:
            wb = _opx.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb["Leads"]
            hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            col_nome   = hdrs.get("Nome",   1)
            col_nicho  = hdrs.get("Nicho",  2)
            col_cidade = hdrs.get("Cidade", 3)
            col_wpp    = hdrs.get("WhatsApp", 11)
            col_email  = hdrs.get("Email",  10)
            col_site   = hdrs.get("Site",    5)
            col_prio   = hdrs.get("Prioridade", 7)

            for row in range(2, min(ws.max_row + 1, 100)):
                nome = ws.cell(row, col_nome).value
                if not nome:
                    continue
                wpp = str(ws.cell(row, col_wpp).value or "")
                email = str(ws.cell(row, col_email).value or "")
                chave = re.sub(r"\D", "", wpp) if wpp else email.lower()
                crm_entry = crm_data.get(chave, {})

                leads.append({
                    "id": chave,
                    "nome": str(nome),
                    "nicho": str(ws.cell(row, col_nicho).value or ""),
                    "cidade": str(ws.cell(row, col_cidade).value or ""),
                    "wpp": wpp,
                    "email": email,
                    "site": str(ws.cell(row, col_site).value or "") if col_site else "",
                    "prioridade": str(ws.cell(row, col_prio).value or ""),
                    "arquivo": path.name,
                    "status": crm_entry.get("status", "prospectado"),
                    "notas": crm_entry.get("notas", ""),
                    "ultima_atualizacao": crm_entry.get("ultima_atualizacao", ""),
                })
            wb.close()
        except Exception:
            pass

    kanban = {
        "prospectado": [],
        "contatado": [],
        "interessado": [],
        "proposta_enviada": [],
        "fechado": [],
        "perdido": [],
    }
    for lead in leads:
        kanban.setdefault(lead["status"], []).append(lead)

    return jsonify({"leads": leads, "kanban": kanban, "total": len(leads)})


@app.route("/api/crm/status", methods=["PATCH"])
def crm_update_status():
    """Atualiza status de um lead no CRM."""
    data = request.json or {}
    lead_id = data.get("id", "")
    status  = data.get("status", "prospectado")
    notas   = data.get("notas", "")

    VALID_STATUS = ["prospectado", "contatado", "interessado", "proposta_enviada", "fechado", "perdido"]
    if status not in VALID_STATUS:
        return jsonify({"error": "Status invalido"}), 400

    crm = _carregar_crm()
    crm[lead_id] = {
        "status": status,
        "notas": notas,
        "ultima_atualizacao": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    _salvar_crm(crm)
    return jsonify({"ok": True})


# ENRIQUECIMENTO
@app.route("/api/enrich/site", methods=["POST"])
def enrich_site():
    """Analisa website: SSL, mobile, velocidade, CMS, problemas."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data = request.json or {}
    url = data.get("url", "")
    if not url:
        return jsonify({"erro": "URL nao informada"}), 400
    try:
        resultado = analisar_website(url)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/enrich/seo", methods=["POST"])
def enrich_seo():
    """Analise SEO basica do site."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data = request.json or {}
    url   = data.get("url", "")
    nome  = data.get("nome", "")
    nicho = data.get("nicho", "")
    if not url:
        return jsonify({"erro": "URL nao informada"}), 400
    try:
        resultado = analisar_seo(url, nome, nicho)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/enrich/redes", methods=["POST"])
def enrich_redes():
    """Busca redes sociais da empresa."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data  = request.json or {}
    nome  = data.get("nome", "")
    site  = data.get("site", "")
    cidade= data.get("cidade", "")
    try:
        resultado = buscar_redes_sociais(nome, site, cidade)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/enrich/emails", methods=["POST"])
def enrich_emails():
    """Extrai emails do site da empresa."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data = request.json or {}
    url  = data.get("url", "")
    if not url:
        return jsonify({"erro": "URL nao informada"}), 400
    try:
        emails = extrair_emails_do_site(url)
        return jsonify({"emails": emails, "total": len(emails)})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/enrich/telefone", methods=["POST"])
def enrich_telefone():
    """Busca informacoes de um numero de telefone."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data   = request.json or {}
    numero = data.get("numero", "")
    if not numero:
        return jsonify({"erro": "Numero nao informado"}), 400
    try:
        resultado = buscar_dono_telefone(numero)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/enrich/cnpj", methods=["POST"])
def enrich_cnpj_route():
    """Busca dados completos de um CNPJ (tenta 3 APIs)."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data = request.json or {}
    cnpj = re.sub(r"\D", "", data.get("cnpj", ""))
    if len(cnpj) != 14:
        return jsonify({"erro": "CNPJ invalido (precisa ter 14 digitos)"}), 400
    try:
        resultado = enriquecer_cnpj(cnpj)
        if not resultado:
            return jsonify({"erro": "CNPJ nao encontrado em nenhuma das APIs"}), 404
        return jsonify(resultado)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/enrich/score", methods=["POST"])
def enrich_score():
    """Calcula score de oportunidade de um lead."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data = request.json or {}
    lead = data.get("lead", {})
    site = lead.get("website", lead.get("site", ""))
    if site:
        lead["_analise_site"] = analisar_website(site)
        lead["_seo"] = analisar_seo(site, lead.get("nome", ""), lead.get("nicho", ""))
    resultado = calcular_score_oportunidade(lead)
    return jsonify(resultado)


@app.route("/api/enrich/completo", methods=["POST"])
def enrich_completo():
    """Enriquecimento completo de um lead."""
    if not ENRICHER_OK:
        return jsonify({"erro": "Modulo enricher nao disponivel"}), 500
    data  = request.json or {}
    lead  = data.get("lead", {})
    opcoes = data.get("opcoes", {
        "site": True, "seo": True, "redes": True, "emails": True, "score": True
    })

    site  = lead.get("website", lead.get("site", ""))
    nome  = lead.get("nome", "")
    nicho = lead.get("nicho", "")
    resultado = dict(lead)

    if opcoes.get("site") and site:
        resultado["analise_site"] = analisar_website(site)

    if opcoes.get("seo") and site:
        resultado["seo"] = analisar_seo(site, nome, nicho)

    if opcoes.get("redes"):
        resultado["redes_sociais"] = buscar_redes_sociais(nome, site)

    if opcoes.get("emails") and site:
        resultado["emails_encontrados"] = extrair_emails_do_site(site)

    if opcoes.get("score"):
        lead_para_score = {**lead}
        if "analise_site" in resultado:
            lead_para_score["_analise_site"] = resultado["analise_site"]
        if "seo" in resultado:
            lead_para_score["_seo"] = resultado["seo"]
        if "redes_sociais" in resultado:
            lead_para_score["_redes"] = resultado["redes_sociais"]
        resultado["score"] = calcular_score_oportunidade(lead_para_score)

    return jsonify(resultado)


# PROPOSTA COMERCIAL
@app.route("/api/proposal/gerar", methods=["POST"])
def proposal_gerar():
    """Gera proposta comercial HTML personalizada para um lead."""
    data     = request.json or {}
    lead     = data.get("lead", {})
    servicos = data.get("servicos", None)

    try:
        from core.proposal import gerar_proposta_html
    except ImportError:
        return jsonify({"erro": "Modulo proposal nao disponivel"}), 500

    autor = {
        "nome":      os.getenv("SEU_NOME",  "Desenvolvedor Web"),
        "email":     os.getenv("SEU_EMAIL", ""),
        "whatsapp":  os.getenv("SEU_WHATSAPP", ""),
        "site":      os.getenv("SEU_LINK",  ""),
    }

    try:
        html = gerar_proposta_html(lead, autor, servicos)
        return jsonify({"html": html, "ok": True})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/proposal/precos", methods=["GET"])
def proposal_precos():
    """Retorna tabela de precos padrao para propostas."""
    try:
        from core.proposal import PRECOS_PADRAO
        return jsonify(PRECOS_PADRAO)
    except ImportError:
        return jsonify({}), 500


# CAPTCHA HANDLING
@app.route("/api/captcha/status", methods=["GET"])
def captcha_status():
    """Verifica se ha CAPTCHA aguardando resolucao."""
    return jsonify({
        "aguardando": _state.get("captcha_pendente", False),
        "msg": _state.get("captcha_msg", "")
    })


@app.route("/api/captcha/resolvido", methods=["POST"])
def captcha_resolvido():
    """Sinaliza que o usuario resolveu o CAPTCHA manualmente."""
    _state["captcha_pendente"] = False
    evt = _state.get("captcha_evento")
    if evt:
        evt.set()
    _log("CAPTCHA resolvido — continuando scraping...", "success")
    return jsonify({"ok": True})


# Entry point
if __name__ == "__main__":
    _ip = _get_local_ip()
    print("=" * 58)
    print("  Leads Scraper Pro — Interface Web")
    print(f"  PC (localhost):    http://localhost:5000")
    print(f"  Celular (Wi-Fi):   http://{_ip}:5000")
    print(f"  Relatorio:         http://{_ip}:5000/relatorio")
    print(f"  Central Leads:     http://{_ip}:5000/central")
    print("=" * 58)

    def _open():
        time.sleep(1.5)
        webbrowser.open("http://localhost:5000")

    threading.Thread(target=_open, daemon=True).start()
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)
